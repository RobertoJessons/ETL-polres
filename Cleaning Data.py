import pandas as pd
import re
import math
import time
import urllib.request
import json
import os
import tkinter as tk
from datetime import datetime
from rapidfuzz import process as fuzz_process, fuzz as fuzz_scorer

# =====================================================================
# FUNGSI PARSING TANGGAL (dipindah ke level modul)
# ---------------------------------------------------------------------
# Dipindah keluar dari jalankan_etl() supaya bisa dipakai langsung oleh
# GUI (main.py / main_ctk.py) tanpa perlu menjalankan seluruh pipeline.
# =====================================================================

bulan_id = {
    'januari': '01', 'februari': '02', 'maret': '03',
    'april': '04', 'mei': '05', 'juni': '06',
    'juli': '07', 'agustus': '08', 'september': '09',
    'oktober': '10', 'november': '11', 'desember': '12'
}
BULAN_RE = '|'.join(bulan_id.keys())


# ===============================
# FUNGSI KONVERSI ANGKA ROMAWI
# ===============================

def romawi_ke_angka(romawi_str):
    romawi_map = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}
    romawi_str = romawi_str.upper().strip()
    if not romawi_str:
        return None
    if not re.match(r'^[IVXLCDM]+$', romawi_str):
        return None
    hasil = 0
    prev = 0
    for char in reversed(romawi_str):
        nilai = romawi_map.get(char, 0)
        if nilai == 0:
            return None
        if nilai < prev:
            hasil -= nilai
        else:
            hasil += nilai
        prev = nilai
    if hasil < 1 or hasil > 3999:
        return None
    return hasil

def normalisasi_romawi_dalam_teks(text):
    if pd.isna(text) or str(text).strip() == '':
        return text
    text = str(text)
    text = re.sub(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', r'\1/\2/\3', text)
    text = re.sub(r'\.', '', text)
    text = re.sub(r'\s*([/\-])\s*', r'\1', text)
    def ganti_romawi(m):
        prefix = m.group(1)
        romawi = m.group(2)
        suffix = m.group(3)
        angka  = romawi_ke_angka(romawi)
        if angka is not None:
            return f"{prefix}{str(angka).zfill(2)}{suffix}"
        return m.group(0)
    pola_romawi_bulan = r'(\d{1,2}[/\-])([IVXLCDMivxlcdm]{1,6})([/\-]\d{2,4})'
    text = re.sub(pola_romawi_bulan, ganti_romawi, text)
    return text

def _ts(tahun, bulan, hari=1):
    try:
        return pd.Timestamp(f"{tahun}-{bulan}-{str(hari).zfill(2)}")
    except Exception:
        return None

def _bersihkan(text):
    t = str(text).replace('\u00a0', ' ').strip().lower()
    return re.sub(r'\s+', ' ', t)

def _parse_tgl_tunggal(text):
    t = _bersihkan(text)
    t = re.sub(r"^(?:senin|selasa|rabu|kamis|jum'?at|sabtu|minggu)\s*,?\s*", '', t)
    t = re.sub(r'^(?:tanggal|tgl)\s+', '', t)
    t = re.sub(r'^(?:sekitar\s+)?(?:bulan\s+)?', '', t)
    t = re.sub(r'^(?:mulai\s+)?(?:bulan\s+)?', '', t)
    t = re.sub(r'^tahun\s+', '', t)
    t = t.strip(' ,')
    m = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$', t)
    if m:
        return _ts(m.group(3), m.group(2), m.group(1))
    m = re.search(r'(\d{1,2})\s+(' + BULAN_RE + r')\s+(\d{4})', t)
    if m:
        return _ts(m.group(3), bulan_id[m.group(2)], m.group(1))
    m = re.search(r'(\d{1,2})[-/](\d{1,2})?\s+(' + BULAN_RE + r')\s+(\d{4})', t)
    if m and m.group(1):
        return _ts(m.group(4), bulan_id[m.group(3)], m.group(1))
    m = re.match(r'^(' + BULAN_RE + r')\s+(\d{4})$', t.strip())
    if m:
        return _ts(m.group(2), bulan_id[m.group(1)], 1)
    m = re.match(r'^(\d{4})$', t.strip())
    if m:
        return _ts(m.group(1), '01', 1)
    return None

def parse_range_tanggal(text):
    if pd.isna(text) or str(text).strip() == '':
        return None, None
    text_str = str(text).strip()
    text_str = normalisasi_romawi_dalam_teks(text_str)
    m_timestamp = re.match(r'^(\d{4})-(\d{2})-(\d{2})(\s+\d{2}:\d{2}:\d{2})?$', text_str)
    if m_timestamp:
        ts = _ts(m_timestamp.group(1), m_timestamp.group(2), m_timestamp.group(3))
        return ts, ts
    t = _bersihkan(text_str)
    m = re.match(r'^(\d{1,2})-(\d{1,2})\s+(' + BULAN_RE + r')\s+(\d{4})', t)
    if m:
        bulan = bulan_id[m.group(3)]
        return _ts(m.group(4), bulan, m.group(1)), _ts(m.group(4), bulan, m.group(2))
    sep = r'\s+(?:s/d|sampai\s+dengan|sampai)\s+'
    bagian = re.split(sep, t, maxsplit=1, flags=re.IGNORECASE)
    if len(bagian) == 2:
        mulai   = _parse_tgl_tunggal(bagian[0])
        selesai = _parse_tgl_tunggal(bagian[1])
        if mulai or selesai:
            return mulai, selesai
    m = re.match(r'^(' + BULAN_RE + r')\s+(\d{4})\s*-\s*(' + BULAN_RE + r')\s+(\d{4})', t)
    if m:
        return _ts(m.group(2), bulan_id[m.group(1)], 1), _ts(m.group(4), bulan_id[m.group(3)], 1)
    if ',' in t:
        bagian = re.split(r',\s*', t)
        if len(bagian) == 2:
            if re.match(r'^(?:senin|selasa|rabu|kamis|jum\'?at|sabtu|minggu)$', bagian[0]):
                t = _bersihkan(text_str)
                tgl = _parse_tgl_tunggal(t)
                return tgl, tgl
            mulai   = _parse_tgl_tunggal(bagian[0])
            selesai = _parse_tgl_tunggal(bagian[1])
            if mulai and selesai and mulai != selesai:
                return mulai, selesai
    tgl = _parse_tgl_tunggal(t)
    return tgl, tgl

def parse_tanggal(text):
    if pd.isna(text):
        return None
    text_str = str(text).replace('\u00a0', ' ').strip()
    try:
        angka = float(text_str)
        if angka == int(angka) and 1 <= int(angka) <= 99999:
            return pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(angka))
    except (ValueError, TypeError):
        pass
    text_str = normalisasi_romawi_dalam_teks(text_str)
    text_lower = text_str.lower().strip()
    text_lower = re.sub(r'^(senin|selasa|rabu|kamis|jumat|sabtu|minggu)\s*,?\s*', '', text_lower).strip()
    match = re.search(r'(\d{1,2})\s+([a-z]+)\s+(\d{4})', text_lower)
    if match:
        hari  = match.group(1).zfill(2)
        bulan = bulan_id.get(match.group(2))
        tahun = match.group(3)
        if bulan:
            try:
                return pd.Timestamp(f"{tahun}-{bulan}-{hari}")
            except Exception:
                return None
    return text_str



# ===============================
# FUNGSI FORMAT TANGGAL
# ===============================

def format_tanggal_output(val):
    if val is None or pd.isna(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val.strftime('%d/%m/%Y')
    if isinstance(val, str):
        if val.strip() == '':
            return None
        val = normalisasi_romawi_dalam_teks(val)
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})(\s+\d{2}:\d{2}:\d{2})?$', val.strip())
        if m:
            try:
                return pd.Timestamp(f"{m.group(1)}-{m.group(2)}-{m.group(3)}").strftime('%d/%m/%Y')
            except Exception:
                pass
        m2 = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', val.strip())
        if m2:
            return f"{m2.group(1).zfill(2)}/{m2.group(2).zfill(2)}/{m2.group(3)}"
        m3 = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', val.strip())
        if m3:
            try:
                return pd.Timestamp(f"{m3.group(3)}-{m3.group(2).zfill(2)}-{m3.group(1).zfill(2)}").strftime('%d/%m/%Y')
            except Exception:
                pass
        ts = parse_tanggal(val)
        if isinstance(ts, pd.Timestamp):
            return ts.strftime('%d/%m/%Y')
        return val
    try:
        return pd.Timestamp(val).strftime('%d/%m/%Y')
    except Exception:
        return None


# =====================================================================
# VALIDASI PASAL LEWAT API PASAL.ID (opsional, cache per KASUS yang sudah
# dibersihkan -- bukan per baris/per pasal, supaya hemat panggilan API
# dan menghindari rate limit)
# ---------------------------------------------------------------------
# Fitur ini OPSIONAL: hanya aktif kalau API key disediakan lewat env var
# PASAL_ID_API_KEY. Kalau tidak ada, kolom validasi_pasal diisi "-" untuk
# semua baris dan pipeline tetap berjalan normal seperti biasa.
# =====================================================================

REFERENSI_PASAL_CACHE_FILE = "referensi_pasal_kasus_cache.json"
_PASAL_API_BASE = "https://pasal.id/api/v1/search"


def _muat_cache_referensi_pasal():
    if os.path.exists(REFERENSI_PASAL_CACHE_FILE):
        try:
            with open(REFERENSI_PASAL_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _simpan_cache_referensi_pasal(cache):
    try:
        with open(REFERENSI_PASAL_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"  Gagal menyimpan cache referensi pasal: {e}")


def _cari_referensi_pasal_api(nama_kasus, api_key, percobaan_maks=3):
    """Cari pasal yang relevan untuk satu nama kasus lewat API Pasal.id.
    Query sengaja ditambah kata 'KUHP' supaya condong ke KUHP (API ini
    mencari ke SELURUH peraturan Indonesia, bukan cuma KUHP)."""
    import urllib.request
    import urllib.parse
    import urllib.error

    query = urllib.parse.quote(f"{nama_kasus} KUHP")
    url = f"{_PASAL_API_BASE}?q={query}"

    for percobaan in range(percobaan_maks):
        try:
            req = urllib.request.Request(
                url,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "User-Agent": "ETL-Polres-Sleman/1.0",
                },
            )
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))

            kandidat = []
            for item in data.get("results", []):
                matching = (item.get("matching_pasals") or "").strip()
                if not matching:
                    continue
                work = item.get("work") or {}
                judul = (work.get("title") or "")
                nomor_pasal = re.findall(r'\d+[A-Za-z]*', matching)
                if not nomor_pasal:
                    continue
                kandidat.append({
                    "judul": judul,
                    "jenis": work.get("type", ""),
                    "nomor_uu": work.get("number", ""),
                    "tahun_uu": work.get("year", ""),
                    "pasal": nomor_pasal,
                    "score": item.get("score", 0),
                    "kuhp": ("hukum pidana" in judul.lower()) or ("kuhp" in judul.lower()),
                })
            # prioritaskan yang benar-benar dari KUHP, baru urutkan skor
            kandidat.sort(key=lambda x: (not x["kuhp"], -x["score"]))
            return kandidat[:5]

        except Exception as e:
            # urllib.error.HTTPError termasuk subclass Exception, jadi tertangkap di sini juga
            kode = getattr(e, "code", None)
            if kode == 429 and percobaan < percobaan_maks - 1:
                time.sleep(2 ** (percobaan + 1))  # backoff: 2, 4, 8 detik
                continue
            print(f"  Gagal mencari referensi pasal untuk kasus '{nama_kasus}': {e}")
            return []
    return []


def bangun_referensi_pasal_dari_kasus(daftar_kasus_unik, api_key):
    """Untuk tiap nama KASUS unik (setelah dibersihkan/majority-vote), cari
    referensi pasal lewat API -- dengan cache lokal permanen supaya kasus
    yang sama tidak pernah di-lookup dua kali di run berikutnya."""
    cache = _muat_cache_referensi_pasal()
    berubah = False
    daftar_bersih = [k for k in daftar_kasus_unik if k and str(k).strip() not in ('', 'null', 'nan', '-')]
    for i, kasus in enumerate(daftar_bersih):
        if kasus in cache:
            continue
        print(f"  Mencari referensi pasal API untuk kasus: '{kasus}' ({i + 1}/{len(daftar_bersih)})")
        cache[kasus] = _cari_referensi_pasal_api(kasus, api_key)
        berubah = True
        time.sleep(0.5)  # jeda sopan antar-request, hindari rate limit
    if berubah:
        _simpan_cache_referensi_pasal(cache)
    return cache


def validasi_pasal_terhadap_referensi(kasus, pasal_teks, referensi_cache):
    """Bandingkan nomor pasal yang tertulis di data dengan referensi yang
    disarankan API untuk jenis kasus itu. Bukan penghapus/pengubah data --
    hanya penanda untuk ditinjau manual."""
    kasus_key = str(kasus).strip() if kasus and not pd.isna(kasus) else ""
    if not kasus_key or kasus_key not in referensi_cache:
        return "-"  # kasus ini belum ada referensinya (API tidak dipakai / belum di-cache)

    daftar_referensi = referensi_cache.get(kasus_key) or []
    pasal_valid_dari_api = set()
    for ref in daftar_referensi:
        for p in ref.get("pasal", []):
            angka = re.match(r'\d+', p)
            if angka:
                pasal_valid_dari_api.add(angka.group())

    if not pasal_valid_dari_api:
        return "TIDAK ADA REFERENSI DARI API"

    pasal_ditemukan = re.findall(r'\b(\d{2,4})\b', str(pasal_teks or ""))
    pasal_ditemukan = [p for p in pasal_ditemukan if not (len(p) == 4 and int(p) >= 1900)]
    if not pasal_ditemukan:
        return "-"

    if any(p in pasal_valid_dari_api for p in pasal_ditemukan):
        return "OK"
    return f"PERLU DITINJAU: pasal {', '.join(pasal_ditemukan)} tidak cocok referensi API untuk kasus '{kasus_key}'"


def jalankan_etl(path_input="Data_TA_proses.xlsx", progress_callback=None,
                  simpan_csv=True, simpan_xlsx=True, peta_rename_output=None):
    # Menjalankan seluruh pipeline ETL data laporan Polres Sleman.
    # path_input   : path file Excel mentah (Data_TA_proses.xlsx atau pengganti sejenis)
    # progress_callback : opsional, dipanggil jika ingin diberi kabar progres dari GUI
    # simpan_csv / simpan_xlsx : True/False, kontrol format mana yang benar-benar ditulis ke disk
    # peta_rename_output : dict opsional {nama_kolom_output_lama: nama_baru}, diterapkan ke
    #                      processed_data sebelum disimpan (mis. hasil rename kolom di GUI)


    # ===============================
    # NORMALISASI TEKS
    # ===============================

    def normalisasi(text):

        text = str(text).lower()
        text = re.sub(r'[.,-]', ' ', text)
        text = re.sub(r'\s+', ' ', text)

        return text


    # ===============================
    # FUNGSI EKSTRAK NOMINAL UANG DARI KOLOM N
    # ===============================

    def ekstrak_nominal(text):
        if pd.isna(text) or str(text).strip() == "":
            return None, None
        text = str(text).strip()
        pola_nominal = r'Rp\.?\s?[\d.,]+'
        match = re.search(pola_nominal, text, flags=re.IGNORECASE)
        if match:
            nominal_mentah = match.group()
            angka_str = re.sub(r'[Rp.\s,]', '', nominal_mentah)
            try:
                nominal_angka = int(angka_str)
            except ValueError:
                nominal_angka = None
            narasi_bersih = re.sub(pola_nominal, '', text, flags=re.IGNORECASE).strip()
            narasi_bersih = re.sub(r'\s+', ' ', narasi_bersih).strip(' ,-')
            narasi_bersih = narasi_bersih if narasi_bersih else None
        else:
            nominal_angka = None
            narasi_bersih = text
        return nominal_angka, narasi_bersih


    # ===============================
    # FUNGSI AMBIL KETERANGAN SEBELUM UU
    # ===============================

    def ambil_sebelum_uu(text):
        if pd.isna(text):
            return ""
        text = str(text)
        hasil = re.split(r'\bUU\b|\bUndang-Undang\b', text, flags=re.IGNORECASE)[0]
        return hasil.strip()


    # ===============================
    # FUNGSI RINGKAS UU DAN PASAL
    # ===============================

    def _ekstrak_referensi_uu_pasal(teks):
        hasil = []
        t = str(teks).strip()
        # dan\s*/\s*atau menoleransi spasi di sekitar garis miring, mis.
        # "dan / atau", "dan/ atau", "dan /atau" -- semua dianggap 1 pemisah
        # segmen pasal yang utuh (lihat juga _SEP_KOMBINASI untuk kasus serupa
        # pada kolom KASUS).
        segmen_list = re.split(
            r'\bdan\s+atau\b|\bdan\s*/\s*atau\b|\bjuncto\b|\bjo\.?\b',
            t, flags=re.IGNORECASE
        )
        for segmen in segmen_list:
            segmen = segmen.strip()
            uu_no    = None
            uu_tahun = None
            m = re.search(r'(?:UU|Undang-Undang)\s*(?:Nomor|No\.?|RI)?\s*(\d+)\s*[/\s]?\s*(?:Tahun)?\s*(\d{4})', segmen, re.IGNORECASE)
            if m:
                uu_no    = m.group(1)
                uu_tahun = m.group(2)
            if not uu_no:
                m = re.search(r'\bUU\s*(?:RI\s*)?(\d+)/(\d{4})\b', segmen, re.IGNORECASE)
                if m:
                    uu_no    = m.group(1)
                    uu_tahun = m.group(2)
            if not uu_no:
                m = re.search(r'\bUU\s*(?:No\.?\s*)?(\d+)/(\d{4})', segmen, re.IGNORECASE)
                if m:
                    uu_no    = m.group(1)
                    uu_tahun = m.group(2)
            pasal_list = []
            for pm in re.finditer(r'\bPasal\s+(\d+)(?:\s+[Aa]yat\s*\([^)]+\))?', segmen, re.IGNORECASE):
                pasal_list.append(pm.group(1))
            for pm in re.finditer(r'\bdan\s+atau\s+(\d{3,4})\b', segmen, re.IGNORECASE):
                if pm.group(1) not in pasal_list:
                    pasal_list.append(pm.group(1))
            if not pasal_list:
                for pm in re.finditer(r'\b(\d{3,4})\b', segmen):
                    angka = pm.group(1)
                    if len(angka) == 4 and int(angka) >= 1900:
                        continue
                    pasal_list.append(angka)
            if uu_no or pasal_list:
                hasil.append({'uu_no': uu_no, 'uu_tahun': uu_tahun, 'pasal_list': pasal_list,
                               'teks_segmen': segmen})
        return hasil


    def ringkas_uu_pasal(text):
        if pd.isna(text) or str(text).strip() == '':
            return None
        t = str(text).strip()
        referensi = _ekstrak_referensi_uu_pasal(t)
        if not referensi:
            return None
        bagian_output = []
        for ref in referensi:
            uu_no     = ref['uu_no']
            uu_tahun  = ref['uu_tahun']
            pasal_list = list(dict.fromkeys(ref['pasal_list']))
            if not pasal_list:
                continue
            pasal_str = ', '.join([f"Pasal {p}" for p in pasal_list])
            if uu_no and uu_tahun:
                bagian_output.append(f"UU {uu_no} Tahun {uu_tahun} {pasal_str}")
            else:
                bagian_output.append(pasal_str)
        if not bagian_output:
            return None
        return ' | '.join(bagian_output)


    # ===============================
    # LOOKUP TABLE PASAL → UU (DIBANGUN DARI DATA)
    # ===============================

    _PASAL_UU_SEED = {
        '262': ('1', '2023'), '307': ('1', '2023'), '332': ('1', '2023'),
        '351': ('1', '2023'), '412': ('1', '2023'), '446': ('1', '2023'),
        '466': ('1', '2023'), '473': ('1', '2023'), '476': ('1', '2023'),
        '477': ('1', '2023'), '482': ('1', '2023'), '486': ('1', '2023'),
        '488': ('1', '2023'), '492': ('1', '2023'), '521': ('1', '2023'),
        '6'  : ('12', '2022'),
        '44' : ('23', '2004'),
        '29' : ('44', '2008'),
        '100': ('20', '2016'), '102': ('20', '2016'), '103': ('20', '2016'),
        '67' : ('27', '2022'),
        '60' : ('5',  '1997'),
        '81' : ('17', '2016'),
    }

    _PASAL_UU_LOOKUP = dict(_PASAL_UU_SEED)

    def bangun_lookup_dari_data(series_pasal_raw):
        nonlocal _PASAL_UU_LOOKUP
        tambahan = 0
        for teks in series_pasal_raw:
            if pd.isna(teks) or str(teks).strip() == '':
                continue
            referensi = _ekstrak_referensi_uu_pasal(str(teks))
            for ref in referensi:
                if ref['uu_no'] and ref['uu_tahun'] and ref['pasal_list']:
                    for pasal in ref['pasal_list']:
                        if pasal not in _PASAL_UU_LOOKUP:
                            _PASAL_UU_LOOKUP[pasal] = (ref['uu_no'], ref['uu_tahun'])
                            tambahan += 1
        print(f"  Lookup table pasal: {len(_PASAL_UU_LOOKUP)} entri "
              f"({len(_PASAL_UU_SEED)} seed + {tambahan} dari data)")
        return _PASAL_UU_LOOKUP

    def lengkapi_uu_dari_lookup(text):
        if pd.isna(text) or str(text).strip() in ('', 'null', '-', 'nan', 'None'):
            return text
        t = str(text).strip()
        if re.search(r'UU\s+\d+\s+Tahun\s+\d{4}', t, re.IGNORECASE):
            return t
        pasal_ditemukan = re.findall(r'\b(\d{2,4})\b', t)
        pasal_ditemukan = [p for p in pasal_ditemukan if not (len(p) == 4 and int(p) >= 1900)]
        if not pasal_ditemukan:
            return t
        uu_no = None
        uu_tahun = None
        for pasal in pasal_ditemukan:
            if pasal in _PASAL_UU_LOOKUP:
                uu_no, uu_tahun = _PASAL_UU_LOOKUP[pasal]
                break
        pasal_str = ', '.join([f"Pasal {p}" for p in pasal_ditemukan])
        if uu_no and uu_tahun:
            return f"UU {uu_no} Tahun {uu_tahun} {pasal_str}"
        else:
            return pasal_str


    # ===============================
    # NORMALISASI KOLOM E (KASUS) DAN PENCOCOKAN E-F
    # ===============================

    _KASUS_KEYWORD_MAP = {
        'pencurian biasa'               : ['476'],
        'pencurian dengan pemberatan'   : ['477'],
        'pencurian dengan kekerasan'    : ['478'],
        'penggelapan'                   : ['486', '488'],
        'penipuan'                      : ['492'],
        'perbuatan curang'              : ['492'],
        'penganiayaan'                  : ['466'],
        'penganiayaan berat'            : ['467'],
        'pemerasan'                     : ['482'],
        'perampasan'                    : ['482'],
        'pengeroyokan'                  : ['262'],
        'perusakan'                     : ['521'],
        'perkosaan'                     : ['473'],
        'persetubuhan terhadap anak'    : ['473'],
        'kejahatan merek'               : ['100', '102', '103'],
        'kekerasan seksual'             : ['6'],
        'kekerasan dalam rumah tangga'  : ['44'],
        'kdrt'                          : ['44'],
        'pornografi'                    : ['29'],
        'perlindungan data pribadi'     : ['67'],
        'kejahatan perkawinan'          : ['412'],
        'pemalsuan'                     : ['392'],
        'psikotropika'                  : ['60'],
        'ite'                           : ['332'],
        'informasi dan transaksi elektronik': ['332'],
    }

    _SINGKATAN_KASUS = {
        'kdrt'    : 'Kekerasan Dalam Rumah Tangga',
        'curat'   : 'Pencurian Dengan Pemberatan',
        'curas'   : 'Pencurian Dengan Kekerasan',
        'curanmor': 'Pencurian Kendaraan Bermotor',
        'curbe'   : 'Pencurian Dengan Pemberatan',
        'curbiasa': 'Pencurian Biasa',
        'ite'     : 'Kejahatan Informasi Dan Transaksi Elektronik',
        'tpks'    : 'Kekerasan Seksual',
        'tpo'     : 'Tindak Pidana Orang',
        'tipikor' : 'Tindak Pidana Korupsi',
        'narkoba' : 'Narkotika',
        'narkotika': 'Narkotika',
        'perkab'  : 'Persetubuhan Terhadap Anak',
        'pdp'     : 'Kejahatan Terkait Perlindungan Data Pribadi',
        'kuhp'    : None,
        'sajam'   : 'Kepemilikan Senjata Tajam',
    }

    # Mapping normalisasi kasus: typo/variasi/deskripsi panjang → nama standar
    _NORMALISASI_KASUS = {
        # Penganiayaan
        'aniaya'                    : 'Penganiayaan',
        'penganiyaan'               : 'Penganiayaan',
        'di duga penganiayaan'      : 'Penganiayaan',
        'dugaan penganiayaan'       : 'Penganiayaan',

        # Pengeroyokan
        'keroyok'                   : 'Pengeroyokan',
        'pengroyokan'               : 'Pengeroyokan',
        'keroyokaniaya'             : 'Pengeroyokan Dan Penganiayaan',
        'pengroyokanpenganiayaan'   : 'Pengeroyokan Dan Penganiayaan',

        # Penggelapan
        'pengelapan'                : 'Penggelapan',
        'tipu gelap'                : 'Penggelapan',
        'penipuanpenggelapan'       : 'Penipuan Dan Penggelapan',
        'penipuanperbuatan curang'  : 'Penipuan',

        # Pencurian
        'pencurian perhiasan emas'  : 'Pencurian',
        'pencurian sepeda motor'    : 'Pencurian Kendaraan Bermotor',
        'pencurian hp'              : 'Pencurian',
        'pencurian di dalam lingkungan keluarga': 'Pencurian',
        'percobaan pencurian'       : 'Pencurian',
        'pencurian secara bersamasama': 'Pencurian',
        'pencurian ringan'          : 'Pencurian Biasa',

        # Kekerasan seksual & cabul
        'perbuatan cabul'           : 'Pencabulan',
        'cabul'                     : 'Pencabulan',
        'pebuatan cabul'            : 'Pencabulan',
        'kekerasan anak'            : 'Kekerasan Terhadap Anak',
        'tindak pidana kekerasan seksual'   : 'Kekerasan Seksual',
        'tindak pidana kekekerasan seksual' : 'Kekerasan Seksual',
        'kekersan seksual'          : 'Kekerasan Seksual',

        # KDRT
        'kekerasan dalam rumah tanggal'     : 'Kekerasan Dalam Rumah Tangga',
        'kekerasan dalam rumahtangga'       : 'Kekerasan Dalam Rumah Tangga',
        'kekerasan fisik dalam rumah tangga': 'Kekerasan Dalam Rumah Tangga',
        'kekerasan psikis dalam rumah tangga': 'Kekerasan Dalam Rumah Tangga',
        'kekerasan didalam rumah tangga'    : 'Kekerasan Dalam Rumah Tangga',

        # Persetubuhan
        'persetubuhan thd anak'             : 'Persetubuhan Terhadap Anak',
        'persetubuhan anak dan perbuatan cabul terhadap anak': 'Persetubuhan Terhadap Anak Dan Pencabulan',
        'persetubuhan dan pencabulan'       : 'Persetubuhan Terhadap Anak Dan Pencabulan',
        'persetubuhan terhadap anak dan perbuatan cabul': 'Persetubuhan Terhadap Anak Dan Pencabulan',
        'pemerkosaan dan atau persetubuhan terhadap anak': 'Perkosaan Dan Atau Persetubuhan Terhadap Anak',
        'perbuatan cabul atau persetubuhan terhadap anak': 'Persetubuhan Terhadap Anak Dan Pencabulan',
        'persetubuhan terhadap anak atau pencabulan terhadap anak': 'Persetubuhan Terhadap Anak Dan Pencabulan',

        # Pemalsuan
        'pemalsuan suratdokumen'    : 'Pemalsuan Surat',
        'pemalsuan merk'            : 'Pemalsuan Merek',
        'pemalsuan merk dan indaksi geografis': 'Pemalsuan Merek',
        'kejahatan pemalsuan merek' : 'Pemalsuan Merek',

        # Pencemaran nama baik / ITE
        'cemar nama baik'           : 'Pencemaran Nama Baik',
        'ket palsu'                 : 'Keterangan Palsu',
        'illegal akses'             : 'Akses Ilegal Sistem Elektronik',
        'akses tanpa hak'           : 'Akses Ilegal Sistem Elektronik',
        'illegal aksesakses sistem elektronik milik orang lain tanpa hak': 'Akses Ilegal Sistem Elektronik',
        'tindak pidana ite atau tindakpidana fitnah': 'Kejahatan ITE',
        'penipuan ite'              : 'Penipuan Melalui ITE',
        'dengan sengaja menyerang kehormatan atau nama baik melalui sistem elektronik': 'Pencemaran Nama Baik',
        'mempertunjukan dokumen elektronik yang bermuatan asusila': 'Pornografi',
        'menyebarluaskan pornografi' : 'Pornografi',
        'pemerasan melalui media elektronik': 'Pemerasan',

        # Data pribadi
        'mengungkapkan data pribadi orang lain tanpa hak': 'Penyalahgunaan Data Pribadi',
        'penggunaan data pribadi tanpa ijin': 'Penyalahgunaan Data Pribadi',
        'penggunaan data pribadi'   : 'Penyalahgunaan Data Pribadi',
        'perlindungan data diri'    : 'Penyalahgunaan Data Pribadi',
        'penipuan dan penggelapan atau perlindungan data pribadi': 'Penipuan Dan Penggelapan',

        # Zina/Pernikahan
        'zina'                      : 'Perzinahan',
        'perzinahaan'               : 'Perzinahan',
        'melakukan hidup bersama'   : 'Perzinahan',
        'pernikahan siri'           : 'Pernikahan Siri',
        'perzinahan dan atau hidup bersama sebagai suami istri.': 'Perzinahan',

        # Sajam
        'membawa sajam'             : 'Kepemilikan Senjata Tajam',
        'penyalahgunaan sajam'      : 'Kepemilikan Senjata Tajam',
        'kepemilikan kepenguasaan senjata tajam dan membawa sajam tanpa ijin': 'Kepemilikan Senjata Tajam',
        'pembakaran dan membawa sajam tanpa izin': 'Pembakaran Dan Kepemilikan Senjata Tajam',

        # Penipuan variasi
        'penipuanperbuatan curang'  : 'Penipuan',
        'penipuan atau penggelangan': 'Penipuan Atau Penggelapan',
        'tindak pidana penipuan da atau pengelapan': 'Penipuan Dan Atau Penggelapan',

        # Pengeroyokan deskripsi panjang
        'secara bersamasama di muka umum melakukan kekerasan terhadap orang atau penganiayaan': 'Pengeroyokan Dan Penganiayaan',
        'kekerasan secara bersamasama di muka umum terhadap orang': 'Pengeroyokan',
        'pengeroyokan atau secara bersamasama melakukan kekerasan terhadap orang': 'Pengeroyokan',
        'dugaan tindak pidana melakukan kekerasan terhadap orang atau barang secara bersamasama di muka umum': 'Pengeroyokan',
        'kekerasan terhadap anak dan atau pengeroyokan': 'Kekerasan Terhadap Anak Dan Pengeroyokan',
        'kekerasan anak atau pengeroyokan': 'Kekerasan Terhadap Anak Dan Pengeroyokan',
        'pengroyokan dan atau tentang perlindungan anak': 'Pengeroyokan Dan Kekerasan Terhadap Anak',

        # Lain-lain
        'pengancaman dan pemerasan'  : 'Pemerasan Dan Pengancaman',
        'dugaan pengancaman'         : 'Pengancaman',
        'pengrusakan'                : 'Perusakan',
        'pengerusakan'               : 'Perusakan',
        'penganiayaan dan pengrusakan': 'Penganiayaan Dan Perusakan',
        'perampasan dan pengerusakan': 'Perampasan Dan Perusakan',
        'pencurian dan atau pengerusakan': 'Pencurian Dan Perusakan',
        'pencurian dan penggelapan atau pengrusakan': 'Pencurian Dan Penggelapan',
        'perusakan terhadap barang dan atau penganiayaan': 'Perusakan Dan Penganiayaan',
        'pengrusakan dan atau setiap orang dilarang melakukan perbuatan yang dapat menimbulkan gangguan fisik dan elektro magnetik terhadap penyelengaraan telekomunikasi': 'Perusakan',
        'penggelapan dalam keluarga'  : 'Penggelapan',
        'tindak pidana penggelapan barang jaminan fidusia': 'Penggelapan Barang Jaminan Fidusia',
        'pemaksaan dengan kekerasan atau ancaman kekerasan': 'Pemaksaan',
        'karena lalainya mengakibatkan luka': 'Kelalaian Yang Mengakibatkan Luka',
        'melarikan atau menyembunyikan anak': 'Penelantaran Anak',
        'tindak pidana setiap orang dilarang menempatkan, membiarkan, melibatkan, menyuruh, melibatkan anak dalam situasi perlakuan salah dan penelantaran sebagaimana dimaksud dalam': 'Penelantaran Anak',
        'dugaan tindak pidana korupsi dana hasil penjualan atau pelepasan hak atas tanah desa': 'Tindak Pidana Korupsi',
        'dugaan tindak pidana korupsi pada paket belanja langganan free wifi padukuhan, komunitas, dan pasar tradisional': 'Tindak Pidana Korupsi',
        'penggelapan dan atau penipuan': 'Penipuan Dan Atau Penggelapan',
        'penggelapan dalam jabatan atau penggelapan': 'Penggelapan Dalam Jabatan',
        'penipuan dan penggelapan atau perlindungan data pribadi': 'Penipuan Dan Penggelapan',
        'penemuan obat mercon'        : 'Kepemilikan Bahan Berbahaya',

        # Typo/variasi yang masih lolos
        'aniaya'                                          : 'Penganiayaan',
        'di duga penganiayaan'                            : 'Penganiayaan',
        'penganiyaan'                                     : 'Penganiayaan',
        'dugaan tindak pidana melakukan kekerasan terhadap orang' : 'Penganiayaan',

        'pengroyokan'                                     : 'Pengeroyokan',
        'pengroyokanpenganiayaan'                         : 'Pengeroyokan',
        'keroyok'                                         : 'Pengeroyokan',
        'keroyokaniaya'                                   : 'Pengeroyokan',
        'kekerasan secara bersamasama di muka umum terhadap orang': 'Pengeroyokan',
        'pengeroyokan terhadap orang barang'              : 'Pengeroyokan',

        'pengelapan'                                      : 'Penggelapan',
        'tipu gelap'                                      : 'Penggelapan',
        'penipuanpenggelapan'                             : 'Penggelapan',
        'penggelapan dalam keluarga'                      : 'Penggelapan',
        'tindak pidana penggelapan barang jaminan fidusia': 'Penggelapan Barang Jaminan Fidusia',

        'pengerusakan'                                    : 'Perusakan',
        'pengrusakan'                                     : 'Perusakan',

        'penipuanperbuatan curang'                        : 'Penipuan',
        'penipuanperbuatan curang'                        : 'Penipuan',
        'penipuan ite'                                    : 'Penipuan',
        'tindak pidana ite'                               : 'Kejahatan ITE',

        'perbuatan cabul'                                 : 'Pencabulan',
        'pebuatan cabul'                                  : 'Pencabulan',
        'cabul'                                           : 'Pencabulan',

        'tindak pidana kekerasan seksual'                 : 'Kekerasan Seksual',
        'tindak pidana kekekerasan seksual'               : 'Kekerasan Seksual',
        'kekersan seksual'                                : 'Kekerasan Seksual',

        'kekerasan anak'                                  : 'Kekerasan Terhadap Anak',
        'kekerasan dalam rumah tanggal'                   : 'Kekerasan Dalam Rumah Tangga',
        'kekerasan dalam rumahtangga'                     : 'Kekerasan Dalam Rumah Tangga',
        'kekerasan didalam rumah tangga'                  : 'Kekerasan Dalam Rumah Tangga',
        'kekerasan fisik dalam rumah tangga'              : 'Kekerasan Dalam Rumah Tangga',
        'kekerasan psikis dalam rumah tangga'             : 'Kekerasan Dalam Rumah Tangga',

        'persetubuhan thd anak'                           : 'Persetubuhan Terhadap Anak',
        'persetubuhan anak'                               : 'Persetubuhan Terhadap Anak',
        'persetubuhan'                                    : 'Persetubuhan Terhadap Anak',

        'zina'                                            : 'Perzinahan',
        'perzinahaan'                                     : 'Perzinahan',
        'melakukan hidup bersama'                         : 'Perzinahan',

        'pemalsuan suratdokumen'                          : 'Pemalsuan Surat',
        'pemalsuan merk'                                  : 'Pemalsuan Merek',
        'kejahatan pemalsuan merek'                       : 'Pemalsuan Merek',

        'illegal akses'                                   : 'Akses Ilegal Sistem Elektronik',
        'akses tanpa hak'                                 : 'Akses Ilegal Sistem Elektronik',
        'illegal aksesakses sistem elektronik milik orang lain tanpa hak': 'Akses Ilegal Sistem Elektronik',

        'mengungkapkan data pribadi orang lain tanpa hak' : 'Penyalahgunaan Data Pribadi',
        'penggunaan data pribadi tanpa ijin'              : 'Penyalahgunaan Data Pribadi',
        'penggunaan data pribadi'                         : 'Penyalahgunaan Data Pribadi',
        'perlindungan data diri'                          : 'Penyalahgunaan Data Pribadi',

        'mempertunjukan dokumen elektronik yang bermuatan asusila': 'Pornografi',
        'menyebarluaskan pornografi'                      : 'Pornografi',

        'cemar nama baik'                                 : 'Pencemaran Nama Baik',
        'dengan sengaja menyerang kehormatan'             : 'Pencemaran Nama Baik',

        'ket palsu'                                       : 'Keterangan Palsu',
        'pemalsuan identitas'                             : 'Pemalsuan Identitas',

        'membawa sajam'                                   : 'Kepemilikan Senjata Tajam',
        'penyalahgunaan sajam'                            : 'Kepemilikan Senjata Tajam',
        'kepemilikan kepenguasaan senjata tajam'          : 'Kepemilikan Senjata Tajam',

        'pencurian perhiasan emas'                        : 'Pencurian',
        'pencurian sepeda motor'                          : 'Pencurian Kendaraan Bermotor',
        'pencurian hp'                                    : 'Pencurian',
        'pencurian ringan'                                : 'Pencurian Biasa',
        'pencurian di dalam lingkungan keluarga'          : 'Pencurian',
        'pencurian secara bersamasama'                    : 'Pencurian',
        'percobaan pencurian'                             : 'Pencurian',
        'pencurian dengan pemberatan.'                    : 'Pencurian Dengan Pemberatan',

        'pemerasan melalui media elektronik'              : 'Pemerasan',
        'pemaksaan dengan kekerasan'                      : 'Pemaksaan',
        'dugaan pengancaman'                              : 'Pengancaman',
        'penarikan barang tanpa hak'                      : 'Perampasan',
        'melarikan'                                       : 'Penelantaran Anak',
        'karena lalainya mengakibatkan luka'              : 'Kelalaian Yang Mengakibatkan Luka',
        'penyerobotan tanah'                              : 'Penyerobotan Tanah',
        'pernikahan siri'                                 : 'Pernikahan Siri',

        'tindak pidana setiap orang dilarang menempatkan, membiarkan, melibatkan, menyuruh, melibatkan anak dalam situasi perlakuan salah': 'Penelantaran Anak',
        'dugaan tindak pidana korupsi dana hasil penjualan': 'Tindak Pidana Korupsi',
        'dugaan tindak pidana korupsi pada paket belanja langganan free wifi padukuhan, komunitas,': 'Tindak Pidana Korupsi',
        '378'                                             : None,
    }

    def normalisasi_teks_kasus(text):
        if pd.isna(text) or str(text).strip() == '':
            return None
        t = str(text).strip()
        t = re.sub(r'[|*#@~^`!;\'"\\/_=+<>{}\[\]\-]+', '', t)
        t = re.sub(r'\s*(?:pasal|ps\.?|psl\.?)\s*\d+.*$', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\bUU\b.*$', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\b(?:Undang-Undang|Undang\s+Undang)\b.*$', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\([^)]{1,20}\)\s*$', '', t)
        t = re.sub(r'\s+', ' ', t).strip()

        t_lower = t.lower().strip()

        # Cek normalisasi kasus (typo, variasi, deskripsi panjang)
        if t_lower in _NORMALISASI_KASUS:
            hasil = _NORMALISASI_KASUS[t_lower]
            return hasil.title() if hasil else None

        # Cek singkatan
        if t_lower in _SINGKATAN_KASUS:
            hasil = _SINGKATAN_KASUS[t_lower]
            return hasil if hasil else None

        # Expand singkatan dalam kurung
        m = re.match(r'^(.+?)\s*\(([^)]+)\)\s*$', t)
        if m:
            nama_panjang = m.group(1).strip()
            singkatan    = m.group(2).strip().lower()
            if singkatan in _SINGKATAN_KASUS and _SINGKATAN_KASUS[singkatan]:
                t = _SINGKATAN_KASUS[singkatan]
            elif singkatan in _NORMALISASI_KASUS and _NORMALISASI_KASUS[singkatan]:
                t = _NORMALISASI_KASUS[singkatan]
            else:
                t = nama_panjang

        t = re.sub(r'\s+', ' ', t).strip()
        if not t:
            return None
        return t.title()

    def cocokkan_kasus_dan_pasal(kasus_text, pasal_text):
        kasus_str = str(kasus_text).lower().strip() if not pd.isna(kasus_text) else ''
        pasal_str = str(pasal_text).strip() if not pd.isna(pasal_text) else ''
        hasil_pasal = ringkas_uu_pasal(pasal_str) if pasal_str else None
        hasil_pasal = lengkapi_uu_dari_lookup(hasil_pasal) if hasil_pasal else None
        if hasil_pasal and re.search(r'UU\s+\d+\s+Tahun\s+\d{4}', str(hasil_pasal)):
            return hasil_pasal
        if kasus_str:
            for keyword, pasal_list in _KASUS_KEYWORD_MAP.items():
                if keyword in kasus_str:
                    for pasal_no in pasal_list:
                        if pasal_no in _PASAL_UU_LOOKUP:
                            uu_no, uu_tahun = _PASAL_UU_LOOKUP[pasal_no]
                            pasal_gabung = ', '.join([f"Pasal {p}" for p in pasal_list])
                            kandidat = f"UU {uu_no} Tahun {uu_tahun} {pasal_gabung}"
                            if hasil_pasal and hasil_pasal != kandidat:
                                return hasil_pasal
                            return kandidat
        return hasil_pasal


    # ===============================
    # NORMALISASI ASAL LAPORAN
    # ===============================

    # Mapping normalisasi prefix dan nama satuan
    _NORMALISASI_ASAL_LAPORAN = {
        # Standarisasi prefix: SEK → POLSEK, tanpa prefix → POLSEK
        # Format: (pattern_regex, output_standar)
    }

    # Daftar nama satuan yang dikenal beserta bentuk standarnya
    _SATUAN_MAP = {
        'dpb'      : 'POLSEK DPB',
        'dpt'      : 'POLSEK DPT',
        'bsm'      : 'POLSEK BSM',
        'mlati'    : 'POLSEK MLATI',
        'tempel'   : 'POLSEK TEMPEL',
        'sleman'   : 'POLSEK SLEMAN',
        'kalasan'  : 'POLSEK KALASAN',
        'prambanan': 'POLSEK PRAMBANAN',
        'berbah'   : 'POLSEK BERBAH',
        'ngemplak' : 'POLSEK NGEMPLAK',
        'ckr'      : 'POLSEK CKR',
        'godean'   : 'POLSEK GODEAN',
        'gamping'  : 'POLSEK GAMPING',
        'moyudan'  : 'POLSEK MOYUDAN',
        'ngaglik'  : 'POLSEK NGAGLIK',
        'turi'     : 'POLSEK TURI',
        'pakem'    : 'POLSEK PAKEM',
        'seyegan'  : 'POLSEK SEYEGAN',
        'bbh'      : 'POLSEK BERBAH',
    }

    def normalisasi_asal_laporan(val):
        """
        Normalisasi kolom asal_laporan:
        - Uppercase dan strip spasi
        - Standarisasi prefix: SEK → POLSEK, nama tanpa prefix → POLSEK
        - Contoh: "SEK DPB" → "POLSEK DPB", "GODEAN" → "POLSEK GODEAN",
                  "DPT" → "POLSEK DPT", "RES SLEMAN" → "RES SLEMAN" (tidak diubah)
        """
        if val is None or pd.isna(val) or str(val).strip() in ('', 'null', 'nan', '-'):
            return val

        t = str(val).strip().upper()
        t = re.sub(r'\s+', ' ', t)

        # Jika diawali "RES " → biarkan apa adanya (satuan Polres)
        if t.startswith('RES '):
            return t

        # Jika diawali "POLSEK " → sudah standar, cek nama satuannya
        if t.startswith('POLSEK '):
            nama = t[7:].strip().lower()
            if nama in _SATUAN_MAP:
                return _SATUAN_MAP[nama]
            return t

        # Jika diawali "SEK " → ganti jadi "POLSEK "
        if t.startswith('SEK '):
            nama = t[4:].strip().lower()
            if nama in _SATUAN_MAP:
                return _SATUAN_MAP[nama]
            return 'POLSEK ' + t[4:].strip()

        # Jika hanya nama satuan tanpa prefix (misal "GODEAN", "DPT", "BSM")
        t_lower = t.lower()
        if t_lower in _SATUAN_MAP:
            return _SATUAN_MAP[t_lower]

        # Default: kembalikan uppercase
        return t


    # ===============================
    # DATABASE WILAYAH INDONESIA (DARI API wilayah.id)
    # ===============================

    CACHE_FILE = "wilayah_indonesia_cache.json"
    API_BASE   = "https://wilayah.id/api"

    def _fetch_json(url):
        try:
            with urllib.request.urlopen(url, timeout=15) as r:
                return json.loads(r.read())
        except Exception as e:
            print(f"  [WARNING] Gagal fetch {url}: {e}")
            return None

    def bangun_database_wilayah():
        if os.path.exists(CACHE_FILE):
            print("Memuat database wilayah dari cache lokal...")
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        print("Mendownload database wilayah Indonesia dari wilayah.id...")
        print("(Proses ini hanya dilakukan sekali, hasilnya akan disimpan di cache)")
        db = {'provinsi': [], 'kabupaten': [], 'kecamatan': [], 'desa': []}
        res = _fetch_json(f"{API_BASE}/provinces.json")
        if res:
            db['provinsi'] = res['data']
            print(f"  Provinsi: {len(db['provinsi'])} data")
        for prov in db['provinsi']:
            res = _fetch_json(f"{API_BASE}/regencies/{prov['code']}.json")
            if res:
                for kab in res['data']:
                    kab['prov_code'] = prov['code']
                    kab['prov_name'] = prov['name']
                db['kabupaten'].extend(res['data'])
        print(f"  Kabupaten/Kota: {len(db['kabupaten'])} data")
        for i, kab in enumerate(db['kabupaten']):
            res = _fetch_json(f"{API_BASE}/districts/{kab['code']}.json")
            if res:
                for kec in res['data']:
                    kec['kab_code']  = kab['code']
                    kec['kab_name']  = kab['name']
                    kec['prov_code'] = kab['prov_code']
                    kec['prov_name'] = kab['prov_name']
                db['kecamatan'].extend(res['data'])
            if (i + 1) % 50 == 0:
                print(f"    ... kecamatan: {i + 1}/{len(db['kabupaten'])} kabupaten diproses")
        print(f"  Kecamatan: {len(db['kecamatan'])} data")
        for i, kec in enumerate(db['kecamatan']):
            res = _fetch_json(f"{API_BASE}/villages/{kec['code']}.json")
            if res:
                for desa in res['data']:
                    desa['kec_code']  = kec['code']
                    desa['kec_name']  = kec['name']
                    desa['kab_code']  = kec['kab_code']
                    desa['kab_name']  = kec['kab_name']
                    desa['prov_code'] = kec['prov_code']
                    desa['prov_name'] = kec['prov_name']
                db['desa'].extend(res['data'])
            if (i + 1) % 100 == 0:
                print(f"    ... desa: {i + 1}/{len(db['kecamatan'])} kecamatan diproses, {len(db['desa'])} desa terkumpul")
        print(f"  Desa/Kelurahan: {len(db['desa'])} data")
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(db, f, ensure_ascii=False)
        print(f"Database wilayah tersimpan di '{CACHE_FILE}'")
        return db

    DB_WILAYAH = bangun_database_wilayah()

    _nama_desa      = [d['name'].lower() for d in DB_WILAYAH['desa']]
    _nama_kecamatan = [d['name'].lower() for d in DB_WILAYAH['kecamatan']]
    _nama_kabupaten = [d['name'].lower() for d in DB_WILAYAH['kabupaten']]
    _nama_provinsi  = [d['name'].lower() for d in DB_WILAYAH['provinsi']]

    def _cari_wilayah_db(teks, nama_list, data_list, threshold=88, min_length=3):
        """
        Cari teks di database wilayah dengan fuzzy matching.
        FIX: pakai index yang dikembalikan langsung oleh rapidfuzz (hasil[2]),
        bukan nama_list.index(matched_name) — karena banyak nama kecamatan/desa
        yang sama persis muncul di kabupaten/provinsi berbeda di seluruh
        Indonesia (mis. "Depok" juga ada di Jawa Barat). Pakai .index() akan
        selalu mengambil kemunculan PERTAMA di list nasional, yang bisa salah
        wilayah walau skor fuzzy-nya benar.
        """
        if not teks or len(str(teks).strip()) < min_length:
            return None
        teks_lower = str(teks).lower().strip()
        hasil = fuzz_process.extractOne(teks_lower, nama_list, scorer=fuzz_scorer.token_set_ratio)
        if hasil and hasil[1] >= threshold:
            matched_obj = data_list[hasil[2]]
            if len(matched_obj.get('name', '')) >= 3:
                return matched_obj
        return None

    # ── Subset wilayah Kabupaten Sleman & DIY (prioritas pencarian) ──
    # Data TKP pada skripsi ini fokus di wilayah hukum Polres Sleman, jadi
    # banyak nama kecamatan/desa yang juga punya nama kembar di kabupaten
    # lain se-Indonesia (mis. "Depok" juga kota di Jawa Barat). Supaya
    # fuzzy matching tidak salah ambil wilayah lain, kita coba cocokkan ke
    # subset Sleman dulu, lalu subset DIY, baru (kalau diizinkan) fallback
    # ke database nasional.
    _kec_sleman_obj   = [d for d in DB_WILAYAH['kecamatan'] if 'sleman' in str(d.get('kab_name', '')).lower()]
    _desa_sleman_obj  = [d for d in DB_WILAYAH['desa'] if 'sleman' in str(d.get('kab_name', '')).lower()]
    _nama_kec_sleman  = [d['name'].lower() for d in _kec_sleman_obj]
    _nama_desa_sleman = [d['name'].lower() for d in _desa_sleman_obj]

    _kab_diy_obj  = [d for d in DB_WILAYAH['kabupaten'] if 'yogyakarta' in str(d.get('prov_name', '')).lower()]
    _kec_diy_obj  = [d for d in DB_WILAYAH['kecamatan'] if 'yogyakarta' in str(d.get('prov_name', '')).lower()]
    _desa_diy_obj = [d for d in DB_WILAYAH['desa'] if 'yogyakarta' in str(d.get('prov_name', '')).lower()]
    _nama_kab_diy  = [d['name'].lower() for d in _kab_diy_obj]
    _nama_kec_diy  = [d['name'].lower() for d in _kec_diy_obj]
    _nama_desa_diy = [d['name'].lower() for d in _desa_diy_obj]

    def _cari_kabupaten(teks, threshold=83, min_length=3, izinkan_nasional=True):
        """Cari kabupaten: prioritaskan subset DIY dulu, baru fallback ke
        database nasional (kalau izinkan_nasional=True)."""
        hasil = _cari_wilayah_db(teks, _nama_kab_diy, _kab_diy_obj, threshold, min_length)
        if hasil:
            return hasil
        if izinkan_nasional:
            return _cari_wilayah_db(teks, _nama_kabupaten, DB_WILAYAH['kabupaten'], threshold, min_length)
        return None

    def _cari_kecamatan(teks, threshold=83, min_length=3, izinkan_nasional=True):
        """Cari kecamatan: prioritaskan subset Kabupaten Sleman, lalu DIY,
        baru fallback ke database nasional (kalau izinkan_nasional=True)."""
        hasil = _cari_wilayah_db(teks, _nama_kec_sleman, _kec_sleman_obj, threshold, min_length)
        if hasil:
            return hasil
        hasil = _cari_wilayah_db(teks, _nama_kec_diy, _kec_diy_obj, threshold, min_length)
        if hasil:
            return hasil
        if izinkan_nasional:
            return _cari_wilayah_db(teks, _nama_kecamatan, DB_WILAYAH['kecamatan'], threshold, min_length)
        return None

    def _cari_desa(teks, threshold=77, min_length=3, izinkan_nasional=True):
        """Cari desa: prioritaskan subset Kabupaten Sleman, lalu DIY, baru
        fallback ke database nasional (kalau izinkan_nasional=True)."""
        hasil = _cari_wilayah_db(teks, _nama_desa_sleman, _desa_sleman_obj, threshold, min_length)
        if hasil:
            return hasil
        hasil = _cari_wilayah_db(teks, _nama_desa_diy, _desa_diy_obj, threshold, min_length)
        if hasil:
            return hasil
        if izinkan_nasional:
            return _cari_wilayah_db(teks, _nama_desa, DB_WILAYAH['desa'], threshold, min_length)
        return None


    # ===============================
    # PARSER ALAMAT TKP FREEFORM (tanpa format "TITIK KOORDINAT")
    # ---------------------------------------------------------------
    # Data TKP asli kebanyakan berupa alamat bebas (freeform) tanpa koma,
    # mis. "Caturtunggal Depok Sleman Yogyakarta". Fungsi lama hanya bisa
    # menangani format "TITIK KOORDINAT lat,lon, Desa, Kec, Kab, Prov" dan
    # nyaris selalu gagal untuk alamat freeform. Parser ini memindai kata
    # dari KANAN ke KIRI (sesuai hierarki alamat Indonesia: ...desa,
    # kecamatan, kabupaten, provinsi), mencoba window 1 kata dulu baru
    # diperlebar, dan hanya mencocokkan ke slot yang masih kosong supaya
    # tidak salah wilayah akibat nama kembar (mis. "Depok" ada di Sleman
    # maupun di Jawa Barat).
    # ===============================

    _STOP_PREFIX_ALAMAT = re.compile(
        r'\b(?:jl\.?|jln\.?|jalan|gg\.?|gang|dsn\.?|dusun|dk\.?|dukuh|kp\.?|kampung|'
        r'ds\.?|desa|kel\.?|kelurahan|kalurahan|kec\.?|kecamatan|kapanewon|'
        r'kab\.?|kabupaten|kota|alamat|ktr\.?|kantor|d\.?i\.?)\b',
        flags=re.IGNORECASE
    )
    _ABBR_MAP_ALAMAT = {
        'diy': 'DI Yogyakarta',
        'yka': 'Yogyakarta',
    }
    # Kata kualifier arah/generik yang sering muncul sebagai bagian nama
    # dusun (mis. "Cibuk Kidul", "Sumber Lor") tapi berisiko fuzzy-match
    # salah ke nama provinsi/kabupaten lain yang kebetulan mengandung kata
    # sama (mis. "Barat" -> "Jawa Barat"). Jangan dicoba berdiri sendiri.
    _STOPWORD_ARAH = {
        'barat', 'timur', 'utara', 'selatan', 'tengah', 'kidul', 'lor',
        'wetan', 'kulon', 'tenggara', 'baru', 'lama', 'dalam', 'luar',
        'atas', 'bawah', 'pusat',
    }

    def _bersihkan_teks_alamat(teks):
        t = str(teks)
        t = t.replace('"', ' ').replace('\u201c', ' ').replace('\u201d', ' ')
        t = re.sub(r'-', ' ', t)
        t = re.sub(r'\bRT\.?\s*/?\s*RW\.?\s*[:.]?\s*\d+\s*/?\s*\d*', ' ', t, flags=re.IGNORECASE)
        t = re.sub(r'\bRT\.?\s*\d+', ' ', t, flags=re.IGNORECASE)
        t = re.sub(r'\bRW\.?\s*\d+', ' ', t, flags=re.IGNORECASE)
        t = re.sub(r'\bNo\.?\s*\d+\w*', ' ', t, flags=re.IGNORECASE)
        t = re.sub(r'\b0\d{8,}\b', ' ', t)
        for singkat, panjang in _ABBR_MAP_ALAMAT.items():
            t = re.sub(rf'\b{singkat}\b', panjang, t, flags=re.IGNORECASE)
        t = _STOP_PREFIX_ALAMAT.sub(' ', t)
        t = re.sub(r'[^\w\s]', ' ', t)   # koma & tanda baca lain jadi spasi
        t = re.sub(r'\s+', ' ', t).strip()
        return t

    def ekstrak_wilayah_freeform(teks):
        if not teks or len(str(teks).strip()) < 3:
            return None, None, None, None
        teks_bersih = _bersihkan_teks_alamat(teks)
        desa = kecamatan = kabupaten = provinsi = None

        def isi_slot(tipe, nama, obj):
            nonlocal desa, kecamatan, kabupaten, provinsi
            if tipe == 'provinsi':
                provinsi = nama
            elif tipe == 'kota':
                kabupaten = re.sub(r'^(?:kabupaten|kota)\s+', '', nama, flags=re.IGNORECASE).strip()
                if provinsi is None and obj.get('prov_name'):
                    provinsi = obj['prov_name']
            elif tipe == 'kecamatan':
                kecamatan = nama
                # Backfill kabupaten & provinsi dari data resmi API (bukan
                # tebakan fuzzy lagi) -- menjamin konsistensi hierarki dan
                # menutupi kasus teks yang tidak menyebut kabupaten sama
                # sekali (semua data TKP tetap dari Polres Sleman).
                if kabupaten is None and obj.get('kab_name'):
                    kabupaten = re.sub(r'^(?:kabupaten|kota)\s+', '', obj['kab_name'], flags=re.IGNORECASE).strip()
                if provinsi is None and obj.get('prov_name'):
                    provinsi = obj['prov_name']
            elif tipe == 'desa':
                desa = nama
                if kecamatan is None and obj.get('kec_name'):
                    kecamatan = obj['kec_name']
                if kabupaten is None and obj.get('kab_name'):
                    kabupaten = re.sub(r'^(?:kabupaten|kota)\s+', '', obj['kab_name'], flags=re.IGNORECASE).strip()
                if provinsi is None and obj.get('prov_name'):
                    provinsi = obj['prov_name']

        def cocok_slot_kosong(seg):
            """Coba cocokkan segmen HANYA ke slot yang masih kosong. Ini
            mencegah tabrakan nama kembar (mis. kata 'Depok' adalah
            kabupaten di Jawa Barat sekaligus kecamatan di Sleman) --
            begitu slot kabupaten sudah terisi, kata itu otomatis dicoba
            ke kecamatan/desa berikutnya.

            KHUSUS untuk parsing TKP: pencarian kabupaten/kecamatan/desa
            DIBATASI ke wilayah DIY saja (izinkan_nasional=False). TKP dari
            Polres Sleman pasti berada di DIY, jadi tidak perlu (dan
            berbahaya) mencari sampai ke database nasional -- kata pendek
            apa pun (mis. "Laksda" pada "Jl. Laksda Adisucipto") punya
            risiko tinggi ke-fuzzy-match ke desa yang benar-benar tidak
            berhubungan di provinsi lain (mis. "Laksa" di Kabupaten Dairi,
            Sumatera Utara) kalau dicari di seluruh ~83 ribu desa se-
            Indonesia. Provinsi tetap dicari nasional karena daftarnya
            cuma ~38 dan risiko salah cocoknya jauh lebih kecil.
            """
            if len(seg) < 3:
                return None, None, None
            seg_lower = seg.lower()
            if provinsi is None:
                m = _cari_wilayah_db(seg_lower, _nama_provinsi, DB_WILAYAH['provinsi'], threshold=83, min_length=3)
                if m:
                    return 'provinsi', m['name'], m
            if kabupaten is None:
                m = _cari_kabupaten(seg_lower, threshold=83, min_length=3, izinkan_nasional=False)
                if m:
                    return 'kota', m['name'], m
            if kecamatan is None:
                m = _cari_kecamatan(seg_lower, threshold=83, min_length=3, izinkan_nasional=False)
                if m:
                    return 'kecamatan', m['name'], m
            if desa is None:
                m = _cari_desa(seg_lower, threshold=77, min_length=3, izinkan_nasional=False)
                if m:
                    return 'desa', m['name'], m
            return None, None, None

        # Scan per kata dari kanan ke kiri (sesuai hierarki alamat
        # Indonesia: ...desa, kecamatan, kabupaten, provinsi). Window 1
        # kata dicoba dulu, baru diperlebar ke 2-3 kata kalau tidak
        # ketemu -- supaya window lebar tidak "menelan" kata tetangga
        # akibat token_set_ratio yang toleran terhadap kata tambahan.
        kata = [k for k in teks_bersih.split() if k.lower() not in _STOPWORD_ARAH]
        n = len(kata)
        i = n - 1
        while i >= 0:
            matched = False
            for lebar in (1, 2, 3):
                awal = i - lebar + 1
                if awal < 0:
                    continue
                seg = ' '.join(kata[awal:i + 1])
                tipe, nama, obj = cocok_slot_kosong(seg)
                if tipe:
                    isi_slot(tipe, nama, obj)
                    i = awal - 1
                    matched = True
                    break
            if not matched:
                i -= 1
        return desa, kecamatan, kabupaten, provinsi


    # ===============================
    # FUNGSI EKSTRAK KOORDINAT, KABUPATEN, PROVINSI DARI KOLOM I
    # ===============================

    def ekstrak_kolom_i(text):
        if pd.isna(text) or str(text).strip() == '':
            return None, None, None, None, None, None
        t = str(text).strip()
        lat, lon = None, None
        m = re.search(r'TITIK\s+KOORDINAT\s+(-?\d+\.\d+)\s*,\s*(-?\d+\.\d+)', t, flags=re.IGNORECASE)
        if m:
            lat = float(m.group(1))
            lon = float(m.group(2))
        desa_hasil = kec_hasil = kab_hasil = prov_hasil = None
        if m:
            teks_setelah_koordinat = t[m.end():].strip(' ,')
            bagian = [b.strip() for b in teks_setelah_koordinat.split(',') if b.strip()]
            if len(bagian) >= 1: desa_hasil = bagian[0]
            if len(bagian) >= 2: kec_hasil = bagian[1]
            if len(bagian) >= 3:
                kab_teks = bagian[2]
                kab_bersih = re.sub(r'^(?:Kabupaten|Kota)\s+', '', kab_teks, flags=re.IGNORECASE).strip()
                kab_hasil = kab_bersih if kab_bersih else kab_teks
            if len(bagian) >= 4:
                prov_teks = bagian[3]
                prov_bersih = re.sub(r'^Di\s+', '', prov_teks, flags=re.IGNORECASE).strip()
                prov_hasil = prov_bersih if prov_bersih else prov_teks
        # ── Fallback: alamat freeform tanpa "TITIK KOORDINAT" ──
        # Ini mencakup HAMPIR SEMUA data TKP asli, karena format tersebut
        # jarang muncul. Kalau desa/kecamatan masih kosong setelah parsing
        # di atas, coba parser freeform yang memindai kata per kata.
        if desa_hasil is None and kec_hasil is None:
            desa_ff, kec_ff, kab_ff, prov_ff = ekstrak_wilayah_freeform(t)
            if desa_hasil is None: desa_hasil = desa_ff
            if kec_hasil is None: kec_hasil = kec_ff
            if kab_hasil is None: kab_hasil = kab_ff
            if prov_hasil is None: prov_hasil = prov_ff
        if kab_hasil is None:
            t_bersih = re.sub(r'TITIK\s+KOORDINAT\s+[^,]*', '', t, flags=re.IGNORECASE).strip(' ,')
            bagian_fallback = [b.strip() for b in t_bersih.split(',') if b.strip()]
            for i, b in enumerate(bagian_fallback):
                b_up = b.upper()
                if 'KABUPATEN' in b_up or b_up.startswith('KAB') or 'KOTA' in b_up:
                    match_kab = _cari_wilayah_db(b, _nama_kabupaten, DB_WILAYAH['kabupaten'], threshold=83)
                    if match_kab:
                        kab_hasil  = match_kab['name']
                        prov_obj   = next((p for p in DB_WILAYAH['provinsi'] if p['code'] == match_kab['prov_code']), None)
                        prov_hasil = prov_obj['name'] if prov_obj else None
                    break
        if prov_hasil is None:
            teks_parts = [p.strip() for p in t.split(',') if p.strip()]
            if teks_parts:
                teks_akhir = teks_parts[-1]
                teks_akhir = re.sub(r'^Di\s+', '', teks_akhir, flags=re.IGNORECASE).strip()
                match_prov = _cari_wilayah_db(teks_akhir, _nama_provinsi, DB_WILAYAH['provinsi'], threshold=83)
                if match_prov:
                    prov_hasil = match_prov['name']
        return lat, lon, desa_hasil, kec_hasil, kab_hasil, prov_hasil


    # ===============================
    # FUNGSI EKSTRAK JENIS KELAMIN DARI KOLOM J DAN K
    # ===============================

    def ekstrak_jenis_kelamin(text):
        if pd.isna(text) or str(text).strip() == '':
            return None
        t = str(text).lower()
        m = re.search(r'jenis\s+kelamin\s*:\s*(laki-laki|perempuan)', t, flags=re.IGNORECASE)
        if m:
            nilai = m.group(1).strip()
            if 'laki' in nilai:
                return 'Laki-laki'
            elif 'perem' in nilai:
                return 'Perempuan'
        if 'perempuan' in t:
            return 'Perempuan'
        elif 'laki' in t and 'laki-laki' in t:
            return 'Laki-laki'
        return None


    # ===============================
    # FUNGSI EKSTRAK UMUR DARI KOLOM J DAN K
    # ===============================

    def ekstrak_umur(text):
        if pd.isna(text) or str(text).strip() == '':
            return None
        t = str(text).lower()
        m = re.search(r'umur\s*:\s*(\d+)\s*(?:tahun)?', t, flags=re.IGNORECASE)
        if m:
            try:
                umur = int(m.group(1))
                if 1 <= umur <= 150:
                    return umur
            except ValueError:
                pass
        return None


    # ===============================
    # DETEKSI TIPE WILAYAH
    # ===============================

    def _deteksi_tipe_wilayah(teks, prov_filter=None):
        if not teks or len(str(teks).strip()) < 3:
            return None, None, None
        teks_lower = str(teks).lower().strip()
        teks_clean = re.sub(r'^(kabupaten|kota|desa|kelurahan|kecamatan)\s+', '', teks_lower, flags=re.IGNORECASE).strip()
        match_prov = _cari_wilayah_db(teks_clean, _nama_provinsi, DB_WILAYAH['provinsi'], threshold=83, min_length=3)
        if match_prov:
            return 'provinsi', match_prov['name'], match_prov
        match_kab = _cari_wilayah_db(teks_clean, _nama_kabupaten, DB_WILAYAH['kabupaten'], threshold=83, min_length=3)
        if match_kab:
            return 'kota', match_kab['name'], match_kab
        match_kec = _cari_kecamatan(teks_clean, threshold=83, min_length=3)
        if match_kec:
            return 'kecamatan', match_kec['name'], match_kec
        match_desa = _cari_desa(teks_clean, threshold=77, min_length=3)
        if match_desa:
            return 'desa', match_desa['name'], match_desa
        return None, None, None


    def ekstrak_kolom_j(text):
        if pd.isna(text) or str(text).strip() == '':
            return None, None, None, None
        t = str(text).strip()
        match_alamat = re.search(r'alamat\s+([^|]+)', t, flags=re.IGNORECASE)
        if not match_alamat:
            return None, None, None, None
        alamat_text = match_alamat.group(1).strip()
        bagian = [b.strip() for b in alamat_text.split(',') if b.strip()]
        if bagian:
            bagian_pertama_clean = re.sub(
                r'^(?:jl\.?|jalan|kp\.?|kampung|rt\.?|rw\.?\s*|no\.?|nomer|nomor|\d+/\d+|\d+)',
                '', bagian[0], flags=re.IGNORECASE).strip()
            bagian_pertama_clean = re.sub(r'\s+', ' ', bagian_pertama_clean).strip()
            if bagian_pertama_clean and len(bagian_pertama_clean) >= 3:
                bagian[0] = bagian_pertama_clean
        desa_k = kecamatan_k = kota_k = provinsi_k = None
        detected = []
        for idx, b in enumerate(bagian):
            if len(str(b).strip()) < 3 or re.match(r'^[\d/\-\.]+$', str(b).strip()):
                continue
            tipe, nama, obj = _deteksi_tipe_wilayah(b)
            detected.append((b, tipe, nama))
            if tipe == 'desa' and desa_k is None:
                desa_k = nama
            elif tipe == 'kecamatan' and kecamatan_k is None:
                kecamatan_k = nama
            elif tipe == 'kota' and kota_k is None:
                nama_clean = re.sub(r'^(?:kabupaten|kota)\s+', '', nama, flags=re.IGNORECASE).strip()
                kota_k = nama_clean if nama_clean else nama
            elif tipe == 'provinsi' and provinsi_k is None:
                nama_clean = str(nama).strip('. ')
                provinsi_k = nama_clean if nama_clean else nama
        used_values = [desa_k, kecamatan_k, kota_k, provinsi_k]
        unused_bagian = [b for b in bagian if b not in used_values and len(str(b).strip()) >= 3]
        for b in unused_bagian:
            if provinsi_k is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'provinsi':
                    provinsi_k = str(nama).strip('. ')
                    continue
            if kota_k is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'kota':
                    nama_clean = re.sub(r'^(?:kabupaten|kota)\s+', '', nama, flags=re.IGNORECASE).strip()
                    kota_k = nama_clean if nama_clean else nama
                    continue
            if kecamatan_k is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'kecamatan':
                    kecamatan_k = nama
                    continue
            if desa_k is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'desa':
                    desa_k = nama
                    continue
            if desa_k is None and kecamatan_k and not kota_k:
                desa_k = b
            elif kota_k is None and desa_k and kecamatan_k:
                nama_clean = re.sub(r'^(?:kabupaten|kota)\s+', '', b, flags=re.IGNORECASE).strip()
                kota_k = nama_clean if nama_clean else b
            elif provinsi_k is None:
                provinsi_k = str(b).strip('. ')
        return desa_k, kecamatan_k, kota_k, provinsi_k


    def ekstrak_kolom_k(text):
        if pd.isna(text) or str(text).strip() == '':
            return None, None, None, None
        t = str(text).strip()
        match_alamat = re.search(r'alamat\s+([^|]+)', t, flags=re.IGNORECASE)
        if not match_alamat:
            return None, None, None, None
        alamat_text = match_alamat.group(1).strip()
        bagian = [b.strip() for b in alamat_text.split(',') if b.strip()]
        if bagian:
            bagian_pertama_clean = re.sub(
                r'^(?:jl\.?|jalan|kp\.?|kampung|rt\.?|rw\.?\s*|no\.?|nomer|nomor|\d+/\d+|-|\.|-/)',
                '', bagian[0], flags=re.IGNORECASE).strip()
            bagian_pertama_clean = re.sub(r'\s+', ' ', bagian_pertama_clean).strip()
            if bagian_pertama_clean and len(bagian_pertama_clean) >= 3:
                bagian[0] = bagian_pertama_clean
        desa_p = kecamatan_p = kota_p = provinsi_p = None
        for idx, b in enumerate(bagian):
            if len(str(b).strip()) < 3 or re.match(r'^[\d/\-\.]+$', str(b).strip()):
                continue
            tipe, nama, obj = _deteksi_tipe_wilayah(b)
            if tipe == 'desa' and desa_p is None:
                desa_p = nama
            elif tipe == 'kecamatan' and kecamatan_p is None:
                kecamatan_p = nama
            elif tipe == 'kota' and kota_p is None:
                nama_clean = re.sub(r'^(?:kabupaten|kota)\s+', '', nama, flags=re.IGNORECASE).strip()
                kota_p = nama_clean if nama_clean else nama
            elif tipe == 'provinsi' and provinsi_p is None:
                nama_clean = str(nama).strip('. ')
                provinsi_p = nama_clean if nama_clean else nama
        used_values = [desa_p, kecamatan_p, kota_p, provinsi_p]
        unused_bagian = [b for b in bagian if b not in used_values and len(str(b).strip()) >= 3]
        for b in unused_bagian:
            if provinsi_p is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'provinsi':
                    provinsi_p = str(nama).strip('. ')
                    continue
            if kota_p is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'kota':
                    nama_clean = re.sub(r'^(?:kabupaten|kota)\s+', '', nama, flags=re.IGNORECASE).strip()
                    kota_p = nama_clean if nama_clean else nama
                    continue
            if kecamatan_p is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'kecamatan':
                    kecamatan_p = nama
                    continue
            if desa_p is None:
                tipe, nama, obj = _deteksi_tipe_wilayah(b)
                if tipe == 'desa':
                    desa_p = nama
                    continue
            if desa_p is None and kecamatan_p and not kota_p:
                desa_p = b
            elif kota_p is None and desa_p and kecamatan_p:
                nama_clean = re.sub(r'^(?:kabupaten|kota)\s+', '', b, flags=re.IGNORECASE).strip()
                kota_p = nama_clean if nama_clean else b
            elif provinsi_p is None:
                provinsi_p = str(b).strip('. ')
        return desa_p, kecamatan_p, kota_p, provinsi_p


    # ===============================
    # FUNGSI EKSTRAK NIK (16 DIGIT)
    # ===============================

    def ekstrak_nik(text):
        """
        Mengekstrak NIK dari teks identitas.
        - Ekstrak kelompok digit berurutan dengan panjang 15 atau 16 digit
        - NIK valid = 16 digit (akan di-highlight merah jika 15 digit)
        - Jika menemukan karakter bukan digit, reset dan mulai ulang
        """
        if pd.isna(text) or str(text).strip() == '':
            return None
        t = str(text)
        i = 0
        while i < len(t):
            if t[i].isdigit():
                digit_terkumpul = []
                j = i
                while j < len(t):
                    if t[j].isdigit():
                        digit_terkumpul.append(t[j])
                        if len(digit_terkumpul) == 16:
                            return ''.join(digit_terkumpul)
                    else:
                        break
                    j += 1
                if len(digit_terkumpul) == 15:
                    return ''.join(digit_terkumpul)
                i = j + 1
            else:
                i += 1
        return None

    def sensor_nik(nik):
        """
        Sensor 6 digit tanggal lahir NIK (digit ke-7 s.d. ke-12: DDMMYY) jadi
        'X' untuk menjaga keamanan data pribadi, sambil tetap menyisakan
        6 digit kode wilayah di depan dan 4 digit nomor urut di belakang
        (masih berguna untuk analisis/BI, mis. sebaran wilayah korban/pelaku).
        Contoh: "3273011503940001" -> "327301XXXXXX0001"

        NIK yang panjangnya BUKAN 16 digit (mis. 15 digit / tidak lengkap)
        dibiarkan apa adanya (tidak disensor) supaya tetap kebaca sebagai
        data tidak valid oleh validasi highlight merah di file XLSX.
        """
        if not nik:
            return nik
        n = str(nik)
        if len(n) == 16:
            return n[:6] + 'XXXXXX' + n[-4:]
        return n

    def sensor_nik_dalam_teks(text):
        """
        Cari & sensor NIK (15-16 digit berurutan) yang muncul di DALAM teks
        bebas -- dipakai untuk kolom "_raw" (korban_raw, pelaku_raw) yang
        menyimpan teks identitas mentah apa adanya. Teks lain di sekitar NIK
        (nama, alamat, dll) tidak diubah, cuma deretan digit NIK-nya yang
        disensor (pola sama seperti sensor_nik(): sisakan 6 digit depan &
        4 digit belakang, tengahnya diganti 'X').
        """
        if pd.isna(text) or str(text).strip() == '':
            return text
        t = str(text)

        def _ganti(m):
            digit = m.group(0)
            return digit[:6] + ('X' * (len(digit) - 10)) + digit[-4:]

        return re.sub(r'(?<!\d)\d{15,16}(?!\d)', _ganti, t)


    # ===============================
    # MEMBACA FILE EXCEL (SEMUA SHEET)
    # ===============================

    KOLOM_MAP = {
        'no'           : ['NO'],
        'nomor_laporan': ['LP'],
        'tanggal'      : ['TGL.LP', 'TGL LP'],
        'jam_laporan'  : ['JAM LP', 'JAM.LP'],
        'lp'           : ['LP'],
        'kasus'        : ['KASUS'],
        'pasal'        : ['PASAL'],
        'tgl_kej'      : ['TGL.KEJ', 'TGL KEJ', 'TANGGAL KEJADIAN'],
        'jam_kej'      : ['JAM KEJ', 'JAM.KEJ'],
        'tkp'          : ['TKP'],
        'kerugian'     : ['KERUGIAN', 'KERUGIAN '],
        'identitas'    : ['KORBAN'],
        'identitas_p'  : ['TERLAPOR'],
        'modus'        : ['MO'],
        'uraian'       : ['URAIAN', 'URAIAN '],
        'keterangan_o' : ['KETERANGAN'],
        'asal_laporan' : ['ASAL LAPORAN'],
    }

    # Header lengkap yang sudah dikenal saat ini (baseline aktual dari Data_TA_proses.xlsx).
    # Dipakai GUI (main.py) untuk mendeteksi kolom baru yang belum ada di sini.
    HEADER_DIKENAL = [
        'NO', 'LP', 'TGL.LP', 'JAM LP', 'KASUS', 'PASAL', 'TGL.KEJ', 'JAM KEJ',
        'TKP', 'KORBAN', 'TERLAPOR', 'MO', 'URAIAN', 'KERUGIAN', 'KETERANGAN',
        'PENYIDIK', 'ASAL LAPORAN',
    ]

    KOLOM_INDEX = {
        'no'           : 0,   # Kolom A
        'nomor_laporan': 1,   # Kolom B
        'tanggal'      : 2,   # Kolom C
        'jam_laporan'  : 3,   # Kolom D
        'lp'           : 1,   # Kolom B
        'kasus'        : 4,   # Kolom E
        'pasal'        : 5,   # Kolom F
        'tgl_kej'      : 6,   # Kolom G
        'jam_kej'      : 7,   # Kolom H
        'tkp'          : 8,   # Kolom I
        'kerugian'     : 13,  # Kolom N
        'identitas'    : 9,   # Kolom J
        'identitas_p'  : 10,  # Kolom K
        'modus'        : 11,  # Kolom L
        'uraian'       : 12,  # Kolom M
        'keterangan_o' : 14,  # Kolom O
        'asal_laporan' : 16,  # Kolom Q
    }

    def cari_header_row(df_sheet):
        for i in range(min(3, len(df_sheet))):
            vals = [str(v).strip().upper() for v in df_sheet.iloc[i] if pd.notna(v)]
            if 'LP' in vals or 'NO' in vals:
                return i
        return 0

    def cari_baris_mulai_data(df_sheet, header_row, headers):
        """
        Cari baris pertama tempat data sesungguhnya dimulai:
        1) Prioritas: baris pertama di bawah header di mana kolom "NO" bernilai 1.
        2) Fallback: tepat satu baris di bawah baris header (kalau kolom NO
           tidak ditemukan, atau nilai 1 tidak terdeteksi).
        """
        idx_no = None
        for i, h in enumerate(headers):
            if str(h).strip().upper() == 'NO':
                idx_no = i
                break

        if idx_no is not None:
            for i in range(header_row + 1, len(df_sheet)):
                if idx_no >= df_sheet.shape[1]:
                    break
                nilai = df_sheet.iat[i, idx_no]
                if pd.notna(nilai):
                    try:
                        if int(float(str(nilai).strip())) == 1:
                            return i
                    except (ValueError, TypeError):
                        pass

        return header_row + 1

    def baca_sheet(df_sheet, nama_sheet):
        header_row  = cari_header_row(df_sheet)
        headers     = [str(v).strip() if pd.notna(v) else '' for v in df_sheet.iloc[header_row]]
        baris_mulai = cari_baris_mulai_data(df_sheet, header_row, headers)

        data = df_sheet.iloc[baris_mulai:].copy()
        data.columns = headers
        data = data.reset_index(drop=True)

        def ambil_kolom(alternatif, idx_fallback=None):
            for nama in alternatif:
                for col in data.columns:
                    if str(col).strip().upper() == nama.upper():
                        return data[col]
            if idx_fallback is not None:
                try:
                    return df_sheet.iloc[baris_mulai:, idx_fallback].reset_index(drop=True)
                except Exception:
                    pass
            return pd.Series([None] * len(data))

        def ambil_kolom_by_index(idx):
            try:
                if idx < df_sheet.shape[1]:
                    return df_sheet.iloc[baris_mulai:, idx].reset_index(drop=True)
                else:
                    return pd.Series([None] * (len(df_sheet) - baris_mulai))
            except Exception:
                return pd.Series([None] * (len(df_sheet) - baris_mulai))

        hasil = pd.DataFrame({
            'no'              : ambil_kolom(KOLOM_MAP['no'], KOLOM_INDEX['no']).values,
            'nomor_laporan'   : ambil_kolom(KOLOM_MAP['nomor_laporan'], KOLOM_INDEX['nomor_laporan']).values,
            'tanggal_raw'     : ambil_kolom(KOLOM_MAP['tanggal'], KOLOM_INDEX['tanggal']).values,
            'jam_laporan_raw' : ambil_kolom(KOLOM_MAP['jam_laporan'], KOLOM_INDEX['jam_laporan']).values,
            'lp'              : ambil_kolom(KOLOM_MAP['lp'], KOLOM_INDEX['lp']).values,
            'kasus_raw'       : ambil_kolom(KOLOM_MAP['kasus'], KOLOM_INDEX['kasus']).values,
            'pasal_raw'       : ambil_kolom(KOLOM_MAP['pasal'], KOLOM_INDEX['pasal']).values,
            'tkp'             : ambil_kolom(KOLOM_MAP['tkp'], KOLOM_INDEX['tkp']).values,
            'kerugian_raw'    : ambil_kolom(KOLOM_MAP['kerugian'], KOLOM_INDEX['kerugian']).values,
            'tgl_kej_raw'     : ambil_kolom(KOLOM_MAP['tgl_kej'], KOLOM_INDEX['tgl_kej']).values,
            'jam_kej_raw'     : ambil_kolom(KOLOM_MAP['jam_kej'], KOLOM_INDEX['jam_kej']).values,
            'identitas'       : ambil_kolom(KOLOM_MAP['identitas'], KOLOM_INDEX['identitas']).values,
            'identitas_p'     : ambil_kolom(KOLOM_MAP['identitas_p'], KOLOM_INDEX['identitas_p']).values,
            'modus_raw'       : ambil_kolom(KOLOM_MAP['modus'], KOLOM_INDEX['modus']).values,
            'uraian_raw'      : ambil_kolom(KOLOM_MAP['uraian'], KOLOM_INDEX['uraian']).values,
            'keterangan_o_raw': ambil_kolom(KOLOM_MAP['keterangan_o'], KOLOM_INDEX['keterangan_o']).values,
            'asal_laporan_raw': ambil_kolom(KOLOM_MAP['asal_laporan'], KOLOM_INDEX['asal_laporan']).values,
        })

        mask = (
            hasil['lp'].notna() & (hasil['lp'].astype(str).str.strip() != '') |
            hasil['tanggal_raw'].notna() & (hasil['tanggal_raw'].astype(str).str.strip() != '')
        )
        hasil = hasil[mask]
        return hasil

    semua_sheet = pd.read_excel(path_input, sheet_name=None, header=None)

    # ===============================
    # URUTKAN SHEET BERDASARKAN BULAN-TAHUN
    # ===============================

    _URUTAN_BULAN = {
        'JANUARI': 1, 'FEBRUARI': 2, 'FEBUARI': 2, 'MARET': 3,
        'APRIL': 4, 'MEI': 5, 'JUNI': 6, 'JULI': 7,
        'AGUSTUS': 8, 'SEPTEMBER': 9, 'OKTOBER': 10,
        'NOVEMBER': 11, 'DESEMBER': 12
    }

    def _urutan_sheet(nama):
        nama_upper = str(nama).upper().strip()
        for bulan, urutan in _URUTAN_BULAN.items():
            if bulan in nama_upper:
                tahun_match = re.search(r'\d{4}', nama_upper)
                tahun = int(tahun_match.group()) if tahun_match else 0
                return (tahun, urutan)
        return (9999, 99)

    semua_sheet = dict(sorted(semua_sheet.items(), key=lambda x: _urutan_sheet(x[0])))
    print("Urutan sheet yang akan diproses:")
    for nama in semua_sheet:
        print(f"  - {nama}")

    df_list = []
    for nama_sheet, df_sheet in semua_sheet.items():
        if df_sheet.dropna(how='all').empty:
            print(f"Sheet '{nama_sheet}': dilewati (kosong)")
            continue
        hasil = baca_sheet(df_sheet, nama_sheet)
        if len(hasil) > 0:
            df_list.append(hasil)
            print(f"Sheet '{nama_sheet}': {len(hasil)} baris")
        else:
            print(f"Sheet '{nama_sheet}': dilewati (tidak ada data)")

    if not df_list:
        raise ValueError("Tidak ada data yang berhasil dibaca dari semua sheet.")

    df = pd.concat(df_list, ignore_index=True)
    print(f"Total seluruh sheet sebelum filter: {len(df)} baris")

    # ===============================
    # FILTER BARIS TANPA DATA TANGGAL DI KOLOM C
    # ===============================
    df = df[df['tanggal_raw'].notna() & (df['tanggal_raw'].astype(str).str.strip() != '')].reset_index(drop=True)
    print(f"Total setelah filter baris tanpa tanggal (kolom C): {len(df)} baris")

    tanggal    = df['tanggal_raw'].apply(parse_tanggal)
    kasus      = df['kasus_raw'].apply(normalisasi_teks_kasus)

    # ===============================
    # NORMALISASI KASUS KOMBINASI → KASUS MAYORITAS
    # ===============================

    # Separator yang menandai kasus kombinasi
    _SEP_KOMBINASI = re.compile(
        # dan\s*/\s*atau menoleransi spasi di sekitar garis miring, mis.
        # "DAN/ ATAU", "DAN /ATAU", "DAN / ATAU" -- semua dianggap 1 pemisah
        # utuh "dan/atau", bukan terpecah jadi sisa "/" atau "Dan" nyangkut.
        # da\s+atau menoleransi typo "da atau" (kurang huruf "n") -> "dan atau".
        r'\s+(?:dan\s+atau|da\s+atau|dan\s*/\s*atau|atau|dan|,)\s+',
        flags=re.IGNORECASE
    )

    def _ekstrak_kasus_tunggal(text):
        """
        Ekstrak daftar kasus tunggal dari teks kasus (termasuk kombinasi).
        Contoh: "Penipuan Dan Atau Penggelapan" → ["Penipuan", "Penggelapan"]
        """
        if not text or pd.isna(text):
            return []
        parts = _SEP_KOMBINASI.split(str(text).strip())
        return [p.strip() for p in parts if p.strip() and len(p.strip()) >= 3]

    def bangun_frekuensi_kasus(series_kasus):
        """
        Hitung frekuensi kemunculan setiap kasus TUNGGAL dari seluruh data.
        Kasus kombinasi dipecah dulu, lalu masing-masing dihitung.
        """
        freq = {}
        for val in series_kasus:
            if val is None or pd.isna(val) or str(val).strip() in ('', 'null', 'nan'):
                continue
            parts = _ekstrak_kasus_tunggal(str(val))
            if len(parts) == 1:
                # Kasus tunggal: hitung langsung
                k = parts[0].lower()
                freq[k] = freq.get(k, 0) + 1
        return freq

    def normalisasi_kombinasi_ke_mayoritas(text, freq_map):
        """
        Untuk kasus kombinasi, ambil kasus dengan frekuensi tertinggi.
        Jika kasus tunggal, kembalikan apa adanya.
        """
        if not text or pd.isna(text) or str(text).strip() in ('', 'null', 'nan'):
            return text

        t = str(text).strip()
        parts = _ekstrak_kasus_tunggal(t)

        # Jika hanya 1 kasus (bukan kombinasi), kembalikan apa adanya
        if len(parts) <= 1:
            return t

        # Cari kasus dengan frekuensi tertinggi di antara komponen kombinasi
        best_kasus = None
        best_freq  = -1

        for part in parts:
            f = freq_map.get(part.lower(), 0)
            if f > best_freq:
                best_freq  = f
                best_kasus = part

        # Jika tidak ada yang ditemukan di freq_map, ambil yang pertama
        return best_kasus if best_kasus else parts[0]


    keterangan = df['pasal_raw'].apply(ambil_sebelum_uu)
    alamat     = df['tkp']
    kolom_n    = df['kerugian_raw']
    identitas  = df['identitas']



    # ===============================
    # FUNGSI NORMALISASI JAM LAPORAN
    # ===============================

    def normalisasi_jam(val):
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return None
        if pd.isna(val):
            return None
        if isinstance(val, pd.Timestamp):
            return val.strftime('%H.%M') + ' WIB'
        val_str = str(val).strip()
        if val_str == '' or val_str.lower() in ('nan', 'none', 'null', '-'):
            return None
        val_str = val_str.replace(',', '.')
        m = re.search(r'(\d{1,2})[\.:](\d{2})', val_str)
        if m:
            jam   = m.group(1).zfill(2)
            menit = m.group(2).zfill(2)
            return f"{jam}.{menit} WIB"
        return None


    # ===============================
    # FUNGSI NORMALISASI JAM KEJADIAN
    # ===============================

    _KATA_PERKIRAAN_JAM = [
        'sekira', 'sekitar', 'kira-kira', 'kira kira', 'kurang lebih',
        'lebih kurang', 'kirakira', 'sekira-kira', 'diperkirakan',
        'sekitar jam', 'sekira jam', 'jam', 'pukul', 'pkl', '±', '+'
    ]

    def normalisasi_jam_kejadian(val):
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return None
        if pd.isna(val):
            return None
        if isinstance(val, pd.Timestamp):
            return val.strftime('%H.%M') + ' WIB'
        val_str = str(val).strip()
        if val_str == '' or val_str.lower() in ('nan', 'none', 'null', '-'):
            return None
        bulan_list = ['januari','februari','maret','april','mei','juni',
                      'juli','agustus','september','oktober','november','desember']
        if any(b in val_str.lower() for b in bulan_list):
            return None
        t = val_str.lower()
        for kata in _KATA_PERKIRAAN_JAM:
            t = re.sub(r'\b' + re.escape(kata) + r'\b', '', t, flags=re.IGNORECASE)
        t = t.strip()
        t = t.replace(',', '.').replace(':', '.')
        t = t.rstrip('.')
        t = re.sub(r'\bwib\b', '', t, flags=re.IGNORECASE).strip()
        m = re.search(r'(\d{1,2})\.(\d{2})', t)
        if m:
            jam   = m.group(1).zfill(2)
            menit = m.group(2).zfill(2)
            return f"{jam}.{menit} WIB"
        m = re.match(r'^\s*(\d{1,2})\s*$', t)
        if m:
            jam = m.group(1).zfill(2)
            return f"{jam}.00 WIB"
        return None


    # ===============================
    # EKSTRAK KOLOM N (NOMINAL)
    # ===============================

    hasil_n      = kolom_n.apply(ekstrak_nominal)
    nominal_uang = hasil_n.apply(lambda x: x[0])

    # ===============================
    # EKSTRAK KOORDINAT, KABUPATEN, PROVINSI DARI KOLOM I
    # ===============================

    hasil_kolom_i = alamat.apply(ekstrak_kolom_i)
    latitude      = hasil_kolom_i.apply(lambda x: x[0])
    longitude     = hasil_kolom_i.apply(lambda x: x[1])
    desa_col      = hasil_kolom_i.apply(lambda x: x[2])
    kecamatan_db  = hasil_kolom_i.apply(lambda x: x[3])
    kabupaten_col = hasil_kolom_i.apply(lambda x: x[4])
    provinsi_col  = hasil_kolom_i.apply(lambda x: x[5])

    # ===============================
    # EKSTRAK TANGGAL KEJADIAN (MULAI & SELESAI)
    # ===============================

    tgl_kej_parsed  = df['tgl_kej_raw'].apply(parse_range_tanggal)
    tgl_kej_mulai   = tgl_kej_parsed.apply(lambda x: x[0])
    tgl_kej_selesai = tgl_kej_parsed.apply(lambda x: x[1])

    # ===============================
    # EKSTRAK DATA DARI KOLOM J
    # ===============================

    hasil_kolom_j   = identitas.apply(ekstrak_kolom_j)
    desa_k          = hasil_kolom_j.apply(lambda x: x[0])
    kecamatan_k     = hasil_kolom_j.apply(lambda x: x[1])
    kota_k          = hasil_kolom_j.apply(lambda x: x[2])
    provinsi_k      = hasil_kolom_j.apply(lambda x: x[3])
    jenis_kelamin_k = identitas.apply(ekstrak_jenis_kelamin)
    umur_k          = identitas.apply(ekstrak_umur)

    # ===============================
    # EKSTRAK DATA DARI KOLOM K
    # ===============================

    def deteksi_dalam_lidik(text):
        if pd.isna(text) or str(text).strip() == '':
            return None
        t = str(text).lower()
        if re.search(r'dalam\s+lidik', t, flags=re.IGNORECASE):
            return 'DALAM LIDIK'
        return None

    if 'identitas_p' in df.columns:
        identitas_p     = df['identitas_p']
        jenis_kelamin_p = identitas_p.apply(ekstrak_jenis_kelamin)
        umur_p          = identitas_p.apply(ekstrak_umur)
        status_p        = identitas_p.apply(deteksi_dalam_lidik)
        hasil_kolom_k   = identitas_p.apply(ekstrak_kolom_k)
        kota_p          = hasil_kolom_k.apply(lambda x: x[2])
        provinsi_p      = hasil_kolom_k.apply(lambda x: x[3])
        kota_p          = kota_p.where(status_p.isna(), status_p)
        provinsi_p      = provinsi_p.where(status_p.isna(), status_p)
    else:
        jenis_kelamin_p = pd.Series([None] * len(df))
        umur_p          = pd.Series([None] * len(df))
        kota_p          = pd.Series([None] * len(df))
        provinsi_p      = pd.Series([None] * len(df))
        status_p        = pd.Series([None] * len(df))

    # ── POST-PROCESSING: Validasi & bersihkan data kolom J dan K ──

    def validasi_dan_bersih_wilayah(row, kolom_prefix='k'):
        desa_key = f'desa_{kolom_prefix}'
        kec_key  = f'kecamatan_{kolom_prefix}'
        kota_key = f'kota_{kolom_prefix}'
        prov_key = f'provinsi_{kolom_prefix}'
        desa = row.get(desa_key)
        kec  = row.get(kec_key)
        kota = row[kota_key]
        prov = row[prov_key]
        if kolom_prefix != 'p':
            noise_patterns = [r'rt/rw', r'no\.', r'hp\s*\.?\s*\d+', r'\d{9,}', r'telp', r'jalan', r'jl\.?', r'kp\.?']
            for field_name, field_val in [(desa_key, desa), (kec_key, kec), (kota_key, kota), (prov_key, prov)]:
                if field_val and isinstance(field_val, str):
                    field_lower = field_val.lower()
                    has_noise = any(re.search(p, field_lower, flags=re.IGNORECASE) for p in noise_patterns)
                    if len(field_val) > 50:
                        has_noise = True
                    if has_noise:
                        if field_name == desa_key: desa = None
                        elif field_name == kec_key: kec = None
                        elif field_name == kota_key: kota = None
                        elif field_name == prov_key: prov = None
        if kolom_prefix != 'p':
            if desa and isinstance(desa, str):
                match = _cari_wilayah_db(desa.lower(), _nama_desa, DB_WILAYAH['desa'], threshold=77, min_length=3)
                desa = match['name'] if match else None
            if kec and isinstance(kec, str):
                match = _cari_wilayah_db(kec.lower(), _nama_kecamatan, DB_WILAYAH['kecamatan'], threshold=83, min_length=3)
                kec = match['name'] if match else None
        if kota and isinstance(kota, str):
            kota_clean = re.sub(r'^(?:kabupaten|kota)\s+', '', kota, flags=re.IGNORECASE).strip()
            match = _cari_wilayah_db(kota_clean.lower(), _nama_kabupaten, DB_WILAYAH['kabupaten'], threshold=83, min_length=3)
            kota = match['name'] if match else None
        if prov and isinstance(prov, str):
            prov_clean = str(prov).strip('. ')
            match = _cari_wilayah_db(prov_clean.lower(), _nama_provinsi, DB_WILAYAH['provinsi'], threshold=83, min_length=3)
            if match:
                prov = match['name']
            else:
                match_kab = _cari_wilayah_db(prov_clean.lower(), _nama_kabupaten, DB_WILAYAH['kabupaten'], threshold=83, min_length=3)
                match_kec = _cari_wilayah_db(prov_clean.lower(), _nama_kecamatan, DB_WILAYAH['kecamatan'], threshold=83, min_length=3)
                if match_kab and not kota:
                    kota = match_kab['name']
                    prov = None
                elif match_kec and not kec:
                    kec = match_kec['name']
                    prov = None
                else:
                    prov = None
        return pd.Series({
            desa_key: desa if kolom_prefix != 'p' else None,
            kec_key : kec  if kolom_prefix != 'p' else None,
            kota_key: kota,
            prov_key: prov
        })

    temp_df_k = pd.DataFrame({
        'desa_k': desa_k, 'kecamatan_k': kecamatan_k,
        'kota_k': kota_k, 'provinsi_k': provinsi_k,
        'kota_p': kota_p, 'provinsi_p': provinsi_p,
    })
    temp_validated_k = temp_df_k.apply(lambda row: validasi_dan_bersih_wilayah(row, 'k'), axis=1)
    desa_k      = temp_validated_k['desa_k']
    kecamatan_k = temp_validated_k['kecamatan_k']
    kota_k      = temp_validated_k['kota_k']
    provinsi_k  = temp_validated_k['provinsi_k']
    temp_validated_p = temp_df_k.apply(lambda row: validasi_dan_bersih_wilayah(row, 'p'), axis=1)
    kota_p     = temp_validated_p['kota_p']
    provinsi_p = temp_validated_p['provinsi_p']


    # ===============================
    # DATASET FINAL
    # ===============================

    print("\nMembangun lookup table pasal dari data...")
    bangun_lookup_dari_data(df['pasal_raw'])

    processed_data = pd.DataFrame({
        "no"                      : df['no'].values,
        "nomor_laporan"           : df['nomor_laporan'].values,
        "tanggal_laporan"         : tanggal.values,
        "jam_laporan"             : df['jam_laporan_raw'].values,
        "kasus"                   : kasus.values,
        "pasal_raw"               : df['pasal_raw'].values,
        "keterangan_pasal"        : keterangan.values,
        "ringkasan_pasal"         : [cocokkan_kasus_dan_pasal(k, p) for k, p in zip(df['kasus_raw'], df['pasal_raw'])],
        "jam_kejadian_raw"        : df['jam_kej_raw'].values,
        "jam_kejadian"            : df['jam_kej_raw'].values,
        "tkp_raw"                 : df['tkp'].values,
        "desa_tkp"                : desa_col.values,
        "kecamatan_tkp"           : kecamatan_db.values,
        "kabupaten_tkp"           : kabupaten_col.values,
        "provinsi_tkp"            : provinsi_col.values,
        "latitude"                : latitude.values,
        "longitude"               : longitude.values,
        "tanggal_kejadian_mulai"  : tgl_kej_mulai.values,
        "tanggal_kejadian_selesai": tgl_kej_selesai.values,
        "korban_raw"              : df['identitas'].apply(sensor_nik_dalam_teks).values,
        "nik_korban"              : df['identitas'].apply(ekstrak_nik).apply(sensor_nik).values,
        "umur_korban"             : umur_k.values,
        "jenis_kelamin_korban"    : jenis_kelamin_k.values,
        "kota_korban"             : kota_k.values,
        "provinsi_korban"         : provinsi_k.values,
        "pelaku_raw"              : df['identitas_p'].apply(sensor_nik_dalam_teks).values,
        "nik_pelaku"              : df['identitas_p'].apply(ekstrak_nik).apply(sensor_nik).values,
        "kota_pelaku"             : kota_p.values,
        "provinsi_pelaku"         : provinsi_p.values,
        "jenis_kelamin_pelaku"    : jenis_kelamin_p.values,
        "umur_pelaku"             : umur_p.values,
        "status_pelaku"           : status_p.values,
        "modus"                   : df['modus_raw'].values,
        "uraian_kejadian"         : df['uraian_raw'].values,
        "kerugian"                : nominal_uang.values,
        "keterangan"              : df['keterangan_o_raw'].values,
        "asal_laporan"            : df['asal_laporan_raw'].values,
    })


    # ===============================
    # PEMISAHAN BARIS: KASUS GANDA & PASAL BERLAPIS
    # ===============================
    # Ada 2 skenario pemisahan:
    #
    # A) KASUS GANDA + PASAL BERLAPIS SEIMBANG
    #    Kolom KASUS berupa kombinasi ("penganiayaan dan pengeroyokan") DAN
    #    kolom PASAL juga berlapis dengan jumlah segmen yang SAMA ("Pasal 123
    #    dan atau Pasal 456"). Dipasangkan berurutan: kasus ke-1 <-> pasal
    #    ke-1, kasus ke-2 <-> pasal ke-2, dst.
    #
    # B) PASAL BERLAPIS SAJA (kasus tunggal, atau jumlah komponen kasus tidak
    #    seimbang dengan jumlah segmen pasal -> tidak jelas cara pasangnya)
    #    Kolom KASUS tetap sama di semua baris pecahan (dipakai apa adanya /
    #    hasil mayoritas), tapi kolom PASAL tetap dipecah per segmen. Ini
    #    supaya tiap nomor pasal jadi baris/variabel sendiri di BI, bukan
    #    tergabung dalam satu sel dipisah " | ".
    #
    # Kalau PASAL cuma 1 segmen (tidak berlapis) -> baris TIDAK dipecah sama
    # sekali, tetap 1 baris seperti hasil olahan sebelumnya.
    #
    # nomor_laporan dan kolom lain (termasuk pasal_raw) tetap sama persis di
    # semua baris pecahan -- yang berubah cuma kasus & ringkasan_pasal.

    print("\nMemeriksa baris dengan kasus ganda dan pasal berlapis...")

    def _pisahkan_kasus_pasal_ganda(kasus_raw, pasal_raw, baris_default):
        """
        Coba pecah 1 baris kasus ganda & pasal berlapis jadi beberapa baris.
        baris_default = dict hasil olahan baris ini apa adanya (dipakai kalau
        pemisahan tidak bisa/tidak perlu dilakukan).
        Return: list of dict override ({'kasus':.., 'keterangan_pasal':..,
                'ringkasan_pasal':..}), minimal 1 elemen. pasal_raw SENGAJA
        tidak ikut dipecah -- tetap teks mentah utuh untuk keperluan audit.
        """
        if pd.isna(pasal_raw):
            return [{}]

        referensi_list = _ekstrak_referensi_uu_pasal(str(pasal_raw))
        # Buang segmen referensi yang tidak punya nomor pasal sama sekali
        # (misalnya cuma menyebut nama UU tanpa nomor pasal)
        referensi_list = [r for r in referensi_list if r['pasal_list']]

        # Satu "segmen" (hasil pisah dan-atau/juncto/jo) bisa saja masih
        # memuat LEBIH DARI 1 nomor pasal sekaligus, misalnya:
        #   - dipisah koma tanpa dan-atau: "Pasal 38, Pasal 55 UU 36/1999"
        #   - dua kutipan lengkap ditulis berurutan tanpa kata sambung sama
        #     sekali: "...Pasal 473 UU 1/2023 ... Pasal 415 UU 1/2023"
        #   - angka polos dipisah "dan": "308 dan 318 KUH Pidana"
        # Pecah tiap segmen seperti itu jadi 1 entri per nomor pasal, dengan
        # UU yang sama (nomor UU yang ditemukan di segmen itu berlaku untuk
        # semua nomor pasal di dalamnya).
        referensi_per_nomor = []
        for ref in referensi_list:
            for nomor_pasal in dict.fromkeys(ref['pasal_list']):  # unik, urut kemunculan
                referensi_per_nomor.append({
                    'uu_no'      : ref['uu_no'],
                    'uu_tahun'   : ref['uu_tahun'],
                    'pasal_list' : [nomor_pasal],
                    'teks_segmen': ref.get('teks_segmen', ''),
                })
        referensi_list = referensi_per_nomor

        # Pasal tidak berlapis (0 atau 1 segmen valid) -> tidak ada yang perlu
        # dipecah sama sekali
        if len(referensi_list) < 2:
            return [{}]

        kasus_parts = []
        if not pd.isna(kasus_raw):
            kasus_teks = str(kasus_raw)
            # Kalau seluruh teks kasus mentah ini sudah dikenali di kamus
            # _NORMALISASI_KASUS/_SINGKATAN_KASUS (typo, ejaan salah seperti
            # "da atau" alih-alih "dan atau"), pakai versi yang sudah
            # dikoreksi kamus itu SEBELUM dipecah -- supaya pemisah "dan
            # atau"/"atau"/"dan" bisa dikenali dengan benar, bukan mepet ke
            # kata sebelumnya (mis. "...Penipuan Da" akibat typo "da atau").
            kasus_key = kasus_teks.strip().lower()
            if kasus_key in _NORMALISASI_KASUS and _NORMALISASI_KASUS[kasus_key]:
                kasus_teks = _NORMALISASI_KASUS[kasus_key]
            elif kasus_key in _SINGKATAN_KASUS and _SINGKATAN_KASUS[kasus_key]:
                kasus_teks = _SINGKATAN_KASUS[kasus_key]
            kasus_parts = _ekstrak_kasus_tunggal(kasus_teks)
            # Jaring pengaman tambahan: buang sisa kata sambung ("dan"/"atau"/
            # "da") yang masih nempel di awal/akhir tiap komponen -- ini bisa
            # terjadi kalau ada typo lain yang belum ada di kamus normalisasi
            kasus_parts = [
                re.sub(r'^(?:dan|atau|da)\s+|\s+(?:dan|atau|da)$', '', p,
                       flags=re.IGNORECASE).strip()
                for p in kasus_parts
            ]
            kasus_parts = [p for p in kasus_parts if len(p) >= 3]

        # Skenario A: jumlah komponen kasus sama persis dengan jumlah segmen
        # pasal -> pasangkan 1-1
        if len(kasus_parts) >= 2 and len(kasus_parts) == len(referensi_list):
            daftar_kasus_per_baris = [normalisasi_teks_kasus(k) for k in kasus_parts]
        else:
            # Skenario B: kasus tidak bisa dipasangkan 1-1 dengan pasal ->
            # kasus dibiarkan sama di semua baris, cuma pasal yang dipecah
            daftar_kasus_per_baris = [baris_default['kasus']] * len(referensi_list)

        hasil_baris = []
        for kasus_bagian, ref in zip(daftar_kasus_per_baris, referensi_list):
            pasal_list = list(dict.fromkeys(ref['pasal_list']))
            pasal_str = ', '.join([f"Pasal {p}" for p in pasal_list])
            if ref['uu_no'] and ref['uu_tahun']:
                ringkasan_bagian = f"UU {ref['uu_no']} Tahun {ref['uu_tahun']} {pasal_str}"
            else:
                ringkasan_bagian = lengkapi_uu_dari_lookup(pasal_str)
            hasil_baris.append({
                'kasus'           : kasus_bagian,
                # deskripsi umum (mis. "Pengeroyokan") tetap dipakai bersama di
                # semua baris pecahan karena biasanya cuma disebut sekali di awal teks
                'keterangan_pasal': baris_default['keterangan_pasal'],
                'ringkasan_pasal' : ringkasan_bagian,
                # pasal_raw SENGAJA tidak dipecah -- tetap teks mentah utuh di
                # semua baris pecahan, karena dipakai untuk pengecekan/audit
                # terhadap sumber data asli.
            })
        return hasil_baris

    baris_baru = []
    for row, kasus_raw_val, pasal_raw_val in zip(
            processed_data.to_dict('records'), df['kasus_raw'], df['pasal_raw']):
        pecahan = _pisahkan_kasus_pasal_ganda(kasus_raw_val, pasal_raw_val, row)
        for override in pecahan:
            baris = dict(row)
            baris.update(override)
            baris_baru.append(baris)

    jumlah_sebelum_pecah = len(processed_data)
    processed_data = pd.DataFrame(baris_baru, columns=processed_data.columns)
    jumlah_setelah_pecah = len(processed_data)
    if jumlah_setelah_pecah > jumlah_sebelum_pecah:
        print(f"  {jumlah_setelah_pecah - jumlah_sebelum_pecah} baris baru ditambahkan dari "
              f"pemisahan kasus ganda & pasal berlapis (total {jumlah_sebelum_pecah} -> "
              f"{jumlah_setelah_pecah} baris)")
    else:
        print("  Tidak ada kasus ganda/pasal berlapis yang bisa dipecah (jumlah komponen "
              "kasus & pasal tidak sama, atau bukan kombinasi).")


    # ===============================
    # NORMALISASI NULL VALUES
    # ===============================

    NULL_STRING_VALUES = {
        '', 'nan', 'none', 'null', 'n/a', 'na', '-', '--', '---',
        'tidak ada', 'tidak diketahui', 'unknown', '#n/a', '#null!',
        'kosong', 'nihil', 'belum ada', 'belum diketahui'
    }

    def normalisasi_null(val):
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(val, float):
            if math.isnan(val) or math.isinf(val):
                return None
        if isinstance(val, str):
            val_stripped = val.strip()
            if not val_stripped:
                return None
            if val_stripped.lower() in NULL_STRING_VALUES:
                return None
            if re.match(r'^[_\-\.\/\\\s,;:!?*#@]+$', val_stripped):
                return None
            return val_stripped
        if isinstance(val, (int, float)):
            return val
        if isinstance(val, pd.Timestamp):
            if pd.isnull(val):
                return None
            return val
        if isinstance(val, list):
            return val if len(val) > 0 else None
        if isinstance(val, dict):
            return val if len(val) > 0 else None
        return val

    def normalisasi_null_dataframe(df, fill_value="-"):
        df_hasil = df.copy()
        for col in df_hasil.columns:
            dtype = df_hasil[col].dtype
            if dtype == 'object':
                df_hasil[col] = df_hasil[col].apply(normalisasi_null)
                df_hasil[col] = df_hasil[col].fillna(fill_value)
            elif dtype in ('float64', 'float32'):
                df_hasil[col] = df_hasil[col].where(pd.notna(df_hasil[col]), fill_value)
            elif str(dtype).startswith('datetime'):
                df_hasil[col] = df_hasil[col].where(pd.notna(df_hasil[col]), fill_value)
            else:
                df_hasil[col] = df_hasil[col].apply(normalisasi_null)
                df_hasil[col] = df_hasil[col].fillna(fill_value)
        return df_hasil

    def laporan_null(df, label="DataFrame"):
        total = len(df)
        print(f"\n{'='*60}")
        print(f"LAPORAN NULL VALUES — {label}")
        print(f"{'='*60}")
        print(f"{'Kolom':<30} {'Null':>8} {'Terisi':>8} {'% Null':>8}")
        print(f"{'-'*60}")
        for col in df.columns:
            null_count = df[col].isna().sum()
            filled     = total - null_count
            pct        = (null_count / total * 100) if total > 0 else 0
            flag       = " ⚠" if pct > 50 else ""
            print(f"  {col:<28} {null_count:>8} {filled:>8} {pct:>7.1f}%{flag}")
        print(f"{'='*60}")
        print(f"  {'TOTAL BARIS':<28} {total:>8}")
        print(f"{'='*60}\n")

    for col in processed_data.columns:
        if processed_data[col].dtype == 'object':
            processed_data[col] = processed_data[col].apply(normalisasi_null)
        else:
            processed_data[col] = processed_data[col].where(pd.notna(processed_data[col]), None)
            processed_data = normalisasi_null_dataframe(processed_data, fill_value="null")


    # ===============================
    # NORMALISASI FORMAT OUTPUT
    # ===============================

    def normalisasi_format_teks(val, tipe='default'):
        if val is None or pd.isna(val) or str(val).strip() in ('', 'null', '-', 'nan'):
            return val
        val_str = str(val).strip()
        if tipe == 'nama':
            return val_str.title()
        elif tipe == 'upper':
            return val_str.upper()
        else:
            return val_str

    kolom_nama  = ['kasus', 'desa_tkp', 'kecamatan_tkp', 'kabupaten_tkp', 'provinsi_tkp',
                   'kota_korban', 'provinsi_korban', 'kota_pelaku', 'provinsi_pelaku']
    kolom_upper = ['jenis_kelamin_korban', 'jenis_kelamin_pelaku', 'status_pelaku']

    for col in kolom_nama:
        if col in processed_data.columns:
            processed_data[col] = processed_data[col].apply(lambda x: normalisasi_format_teks(x, 'nama'))
    for col in kolom_upper:
        if col in processed_data.columns:
            processed_data[col] = processed_data[col].apply(lambda x: normalisasi_format_teks(x, 'upper'))


    # ===============================
    # FORMAT KOLOM SEBELUM SIMPAN
    # ===============================

    def bersihkan_spasi(val):
        if val is None or pd.isna(val) or str(val).strip() in ('', 'null', 'nan', '-'):
            return val
        return re.sub(r'\s+', ' ', str(val)).strip()

    # Forward fill kolom asal_laporan
    processed_data["asal_laporan"] = processed_data["asal_laporan"].replace('', None)
    processed_data["asal_laporan"] = processed_data["asal_laporan"].replace('null', None)
    processed_data["asal_laporan"] = processed_data["asal_laporan"].ffill()

    # Normalisasi asal_laporan
    processed_data["asal_laporan"] = processed_data["asal_laporan"].apply(normalisasi_asal_laporan)

    processed_data["nomor_laporan"] = processed_data["nomor_laporan"].apply(bersihkan_spasi)
    processed_data["kasus"]         = processed_data["kasus"].apply(bersihkan_spasi)

    # Bangun frekuensi dari kasus tunggal lalu normalisasi kombinasi
    print("\nMembangun frekuensi kasus untuk normalisasi kombinasi...")
    _freq_kasus = bangun_frekuensi_kasus(processed_data["kasus"])
    print(f"  Total kasus tunggal unik: {len(_freq_kasus)}")
    processed_data["kasus"] = processed_data["kasus"].apply(
        lambda x: normalisasi_kombinasi_ke_mayoritas(x, _freq_kasus)
    )

    # ===============================
    # VALIDASI PASAL LEWAT API (OPSIONAL)
    # Dijalankan di sini karena butuh kolom "kasus" yang SUDAH bersih
    # (hasil majority-vote di atas), bukan kasus_raw yang masih berantakan.
    # ===============================
    pasal_api_key = os.environ.get("PASAL_ID_API_KEY")
    if pasal_api_key:
        print("\nMemvalidasi pasal terhadap referensi API Pasal.id (per jenis kasus, sudah di-cache)...")
        kasus_unik = processed_data["kasus"].dropna().unique().tolist()
        referensi_pasal_kasus = bangun_referensi_pasal_dari_kasus(kasus_unik, pasal_api_key)
        processed_data["validasi_pasal"] = [
            validasi_pasal_terhadap_referensi(k, p, referensi_pasal_kasus)
            for k, p in zip(processed_data["kasus"], processed_data["pasal_raw"])
        ]
        jumlah_perlu_tinjau = processed_data["validasi_pasal"].str.startswith("PERLU DITINJAU").sum()
        print(f"  Validasi pasal selesai: {jumlah_perlu_tinjau} baris perlu ditinjau dari {len(processed_data)} baris")
    else:
        print("\n[INFO] PASAL_ID_API_KEY tidak diset -- validasi pasal via API dilewati (kolom validasi_pasal diisi '-').")
        processed_data["validasi_pasal"] = "-"

    processed_data["modus"]         = processed_data["modus"].apply(
        lambda x: re.sub(r'[\s\W]+$', '', re.sub(r'\s+', ' ', str(x).lower().strip()))
        if x is not None and not pd.isna(x) and str(x).strip() not in ('', 'null', 'nan', '-') else x
    )

    processed_data["tanggal_laporan"]      = processed_data["tanggal_laporan"].apply(format_tanggal_output)
    processed_data["jam_laporan"]          = processed_data["jam_laporan"].apply(normalisasi_jam)
    processed_data["jam_kejadian"]         = processed_data["jam_kejadian"].apply(normalisasi_jam_kejadian)
    processed_data["tanggal_kejadian_mulai"]   = processed_data["tanggal_kejadian_mulai"].apply(format_tanggal_output)
    processed_data["tanggal_kejadian_selesai"] = processed_data["tanggal_kejadian_selesai"].apply(format_tanggal_output)


    # ===============================
    # VERIFIKASI DATA SEBELUM SIMPAN
    # ===============================
    print("\n" + "="*120)
    print("VERIFIKASI HASIL EKSTRAKSI DATA WILAYAH")
    print("="*120)
    print(f"\nTotal baris data: {len(processed_data)}")

    print("\nKelengkapan Data Wilayah:")
    print(f"  Desa terisi: {(processed_data['desa_tkp'].notna() & (processed_data['desa_tkp'] != '')).sum()} / {len(processed_data)}")
    print(f"  Kecamatan terisi: {(processed_data['kecamatan_tkp'].notna() & (processed_data['kecamatan_tkp'] != '')).sum()} / {len(processed_data)}")
    print(f"  Kabupaten terisi: {(processed_data['kabupaten_tkp'].notna() & (processed_data['kabupaten_tkp'] != '')).sum()} / {len(processed_data)}")
    print(f"  Provinsi terisi: {(processed_data['provinsi_tkp'].notna() & (processed_data['provinsi_tkp'] != '')).sum()} / {len(processed_data)}")

    print("\nKelengkapan Data Identitas Pelaku/Korban (dari Kolom J):")
    print(f"  Kota_korban terisi: {(processed_data['kota_korban'].notna() & (processed_data['kota_korban'] != '')).sum()} / {len(processed_data)}")
    print(f"  Provinsi_korban terisi: {(processed_data['provinsi_korban'].notna() & (processed_data['provinsi_korban'] != '')).sum()} / {len(processed_data)}")
    print(f"  Jenis Kelamin_korban terisi: {processed_data['jenis_kelamin_korban'].notna().sum()} / {len(processed_data)}")
    print(f"  Umur_korban terisi: {processed_data['umur_korban'].notna().sum()} / {len(processed_data)}")

    print("\nKelengkapan Data Identitas Orang Kedua (dari Kolom K):")
    print(f"  Kota_pelaku terisi: {(processed_data['kota_pelaku'].notna() & (processed_data['kota_pelaku'] != '')).sum()} / {len(processed_data)}")
    print(f"  Provinsi_pelaku terisi: {(processed_data['provinsi_pelaku'].notna() & (processed_data['provinsi_pelaku'] != '')).sum()} / {len(processed_data)}")
    print(f"  Jenis Kelamin_pelaku terisi: {processed_data['jenis_kelamin_pelaku'].notna().sum()} / {len(processed_data)}")
    print(f"  Umur_pelaku terisi: {processed_data['umur_pelaku'].notna().sum()} / {len(processed_data)}")

    desa_list     = processed_data['desa_tkp'].dropna().unique()
    fragments     = ['jok', 'tu', 'ban', 'lango', 'uning', 'angan', 'rongoh', 'daon', 'nya', 'kabu']
    desa_fragments = [d for d in desa_list if d.lower() in fragments]

    print("\n" + "="*120)
    print("SAMPEL DATA KOLOM I (Baris 0-10)")
    print("="*120)
    sample_cols = ['nomor_laporan', 'tanggal_laporan', 'desa_tkp', 'kecamatan_tkp', 'kabupaten_tkp', 'provinsi_tkp', 'latitude', 'longitude']
    print(processed_data[sample_cols].head(10).to_string())

    print("\n" + "="*120)
    print("SAMPEL DATA KOLOM J - IDENTITAS PELAKU/KORBAN (Baris 0-10)")
    print("="*120)
    print(processed_data[['kota_korban', 'provinsi_korban', 'jenis_kelamin_korban', 'umur_korban']].head(10).to_string())

    print("\n" + "="*120)
    print("SAMPEL DATA KOLOM K - IDENTITAS ORANG KEDUA (Baris 0-10)")
    print("="*120)
    print(processed_data[['kota_pelaku', 'provinsi_pelaku', 'jenis_kelamin_pelaku', 'umur_pelaku']].head(10).to_string())


    # ===============================
    # FUNGSI VALIDASI FORMAT KOLOM
    # ===============================

    def _validasi_tanggal_laporan(val):
        if val is None or str(val).strip() in ('', 'null', 'nan', '-'):
            return True
        return bool(re.match(r'^\d{2}/\d{2}/\d{4}$', str(val).strip()))

    def _validasi_jam_laporan(val):
        if val is None or str(val).strip() in ('', 'null', 'nan', '-'):
            return True
        return bool(re.match(r'^\d{2}\.\d{2}\s+WIB$', str(val).strip()))

    def _validasi_ringkasan_pasal(val):
        if val is None or str(val).strip() in ('', 'null', 'nan', '-'):
            return True
        v = str(val).strip()
        pola_lengkap = r'UU\s+\d+\s+Tahun\s+\d{4}\s+Pasal\s+\d+'
        pola_pasal   = r'^Pasal\s+\d+'
        return bool(re.search(pola_lengkap, v, re.IGNORECASE) or re.match(pola_pasal, v, re.IGNORECASE))

    _VALIDASI_KOLOM = {
        'tanggal_laporan' : _validasi_tanggal_laporan,
        'jam_laporan'     : _validasi_jam_laporan,
        'ringkasan_pasal' : _validasi_ringkasan_pasal,
        'jam_kejadian'    : _validasi_jam_laporan,
        'nik_korban'      : lambda x: bool(re.match(r'^\d{6}X{6}\d{4}$', str(x))) if x not in (None, 'null', 'nan', '-', '') else True,
        'validasi_pasal'  : lambda x: not str(x).startswith("PERLU DITINJAU"),
    }


    # ===============================
    # SIMPAN KE EXCEL TERSTRUKTUR
    # ===============================

    def simpan_ke_excel_terstruktur(dataframe, nama_file, nama_sheet="Data"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("ERROR: openpyxl tidak terinstal. Silakan install dengan: pip install openpyxl")
            return False

        wb = Workbook()
        ws = wb.active
        ws.title = nama_sheet

        header_fill      = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font      = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alt_fill         = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        normal_fill      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        error_fill       = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        normal_font      = Font(size=10)
        error_font       = Font(size=10, bold=True, color="FFFFFF")
        normal_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border      = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        col_names = list(dataframe.columns)

        for col_idx, column_title in enumerate(dataframe.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_title)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_alignment
            cell.border    = thin_border

        error_count = 0
        for row_idx, row_data in enumerate(dataframe_to_rows(dataframe, index=False, header=False), 2):
            is_alternate_row = (row_idx % 2 == 0)
            row_fill = alt_fill if is_alternate_row else normal_fill

            for col_idx, cell_value in enumerate(row_data, 1):
                display_value = "null" if (cell_value is None or str(cell_value).strip() == "" or str(cell_value).lower() == "nan") else cell_value
                cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
                col_name = col_names[col_idx - 1]
                if col_name in _VALIDASI_KOLOM and display_value != "null":
                    is_valid = _VALIDASI_KOLOM[col_name](display_value)
                    if not is_valid:
                        cell.fill  = error_fill
                        cell.font  = error_font
                        error_count += 1
                    else:
                        cell.fill = row_fill
                        cell.font = normal_font
                else:
                    cell.fill = row_fill
                    cell.font = normal_font
                cell.alignment = normal_alignment
                cell.border    = thin_border

        print(f"  Total cell tidak sesuai format (highlight merah): {error_count}")

        for col_idx, column_title in enumerate(dataframe.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(column_title))
            for row_idx in range(2, min(102, len(dataframe) + 2)):
                try:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    max_length = max(max_length, len(str(cell_value)) if cell_value else 0)
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        ws.freeze_panes = "A2"
        wb.save(nama_file)
        return True


    # ===============================
    # TERAPKAN RENAME HEADER OUTPUT (kalau ada, mis. dari GUI)
    # ===============================

    if peta_rename_output:
        cocok = {k: v for k, v in peta_rename_output.items() if k in processed_data.columns}
        if cocok:
            processed_data = processed_data.rename(columns=cocok)
            print(f"\nHeader output disesuaikan: {cocok}")

    # ===============================
    # MEMBUAT TIMESTAMP & NAMA FILE OUTPUT
    # Format nama: {namafile_asal}_bersih_{timestamp}
    # ===============================

    timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    nama_dasar = os.path.splitext(os.path.basename(path_input))[0]

    csv_file = None
    xlsx_file = None

    # ===============================
    # SIMPAN KE CSV
    # ===============================

    if simpan_csv:
        try:
            csv_file = f"{nama_dasar}_bersih_{timestamp}.csv"
            processed_data.to_csv(
                csv_file,
                index=False,
                encoding="utf-8-sig"
            )

            print(f"\nDataset tersimpan di '{csv_file}'")

        except Exception as e:
            csv_file = None
            print(f"\nError saat menyimpan file CSV: {e}")

    # ===============================
    # SIMPAN KE XLSX
    # ===============================

    if simpan_xlsx:
        print("\nMembuat file Excel dengan format tabel terstruktur...")

        try:
            xlsx_file = f"{nama_dasar}_bersih_{timestamp}.xlsx"

            if simpan_ke_excel_terstruktur(
                processed_data,
                xlsx_file,
                nama_sheet="Data Laporan"
            ):
                print(f"Dataset tersimpan di '{xlsx_file}'")
            else:
                xlsx_file = None
                print("Error: Gagal membuat file Excel")

        except Exception as e:
            xlsx_file = None
            print(f"Error saat membuat file Excel: {e}")
            print("Pastikan library openpyxl sudah terinstal: pip install openpyxl")

    # ===============================
    # PROSES SELESAI
    # ===============================

    print("\n" + "=" * 120)
    print("PROSES SELESAI")
    print("=" * 120)

    # ===============================
    # VERIFIKASI RINGKAS (langsung dari processed_data di memori,
    # tidak lagi bergantung pada file CSV yang mungkin tidak ditulis)
    # ===============================

    kolom_kasus = "kasus" if "kasus" in processed_data.columns else None
    if kolom_kasus:
        print("\nDistribusi Kasus:")
        print(processed_data[kolom_kasus].value_counts().to_string())

    return {
        "csv_file": csv_file,
        "xlsx_file": xlsx_file,
        "processed_data": processed_data,
    }


if __name__ == "__main__":
    jalankan_etl()
